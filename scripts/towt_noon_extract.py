"""Reprise d'historique TOWT — extraction LOCALE des noon reports (prototype).

À exécuter sur un poste où la bibliothèque « Service Technique / 10 - Data
reporting Noon reports » est synchronisée. Chaque rapport est un classeur Excel
« <NAVIRE> - <Noon|Departure|Arrival> [Report] - AAAA-MM-JJ.xlsx » dont l'onglet
« Reporting form » est un FORMULAIRE (libellés en colonne A, valeurs à droite),
en deux générations connues :
- « Version 3 » (08-09/2024) : 8 colonnes, engins en heures/consommation ;
- « CFOTE_05 Noon Report Rev 2.1 » (dès 09/2024) : 14-16 colonnes, compteurs
  D / D-1 / départ, conso en litres et tonnes, densité, EU-MRV, tirants d'eau.

Le parseur est PILOTÉ PAR LES LIBELLÉS (pas par les adresses de cellules) : il
tolère les lignes insérées et les variantes d'intitulé. Il produit :
- ``noon_reports.ndjson`` : un objet JSON par rapport (toutes les rubriques,
  valeurs brutes conservées à côté des valeurs normalisées, SHA-256 du fichier) ;
- ``noon_reports_summary.csv`` : une ligne par rapport (clés de rapprochement :
  navire, voyage TOWT, type, date/heure UTC, position décimale, conso, ROB).

Il n'écrit RIEN en base : la table de destination (archive immuable) est une
décision d'architecture en attente (ADR-014, décision 6). Dépendance : openpyxl.

Usage :
    python scripts/towt_noon_extract.py --source "<dossier 10 - Data reporting Noon reports>" --out ./noon_towt
    python scripts/towt_noon_extract.py --source ... --out ... --vessel ANEMOS
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - dépendance de l'environnement local
    openpyxl = None

FILE_RE = re.compile(
    r"^(?P<vessel>[A-Za-z_0-9]+)\s*-\s*(?P<kind>Noon|Departure|Arrival)(?:\s+Report)?\s*-\s*"
    r"(?P<date>\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)
FILE_RE_COMPACT = re.compile(
    r"^(?P<vessel>[A-Za-z]+)-(?P<kind>Noon|Departure|Arrival)-(?P<date>\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)
SLOTS = ("16:00", "20:00", "00:00", "04:00", "08:00", "12:00")
ENGINES = (
    "Port Main Engine",
    "Starboard Main Engine",
    "FWD Generator",
    "AFT Generator",
    "Port Shaft generator",
    "PS Shaft generator",
    "Starboard Shaft Generator",
    "STBD Shaft Generator",
)
HOLDS = (
    "Sea water",
    "Air",
    "Cellar",
    "Upper FWD hold",
    "Middle FWD hold",
    "Lower FWD hold",
    "Upper Aft hold",
    "Middle Aft hold",
    "Lower Aft hold",
)
NUM_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def _norm(label: Any) -> str:
    return re.sub(r"[\s:()\[\]°'%]+", " ", str(label or "")).strip().lower()


def _num(value: Any) -> float | None:
    """« 8,3 kt » → 8.3 ; « 0,257 MT » → 0.257 ; « 00 % » → 0.0 ; None si vide."""
    if value is None or value == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    m = NUM_RE.search(str(value))
    return float(m.group(0).replace(",", ".")) if m else None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    s = str(value).strip()
    return s or None


def _tz_offset(value: Any) -> int | None:
    """« UTC-4 » → -4, « UTC » → 0, « UTC+2 » → 2."""
    s = str(value or "").upper().replace(" ", "")
    m = re.fullmatch(r"UTC([+-]\d{1,2})?", s)
    if not m:
        return None
    return int(m.group(1)) if m.group(1) else 0


class FormGrid:
    """Grille du formulaire : recherche par libellé, lecture relative."""

    def __init__(self, rows: list[list[Any]]):
        self.rows = rows

    def find(self, label: str, *, col: int = 0, start: int = 0) -> int | None:
        """Ligne du libellé : correspondance exacte d'abord, préfixe en repli
        (« Time » ne doit pas attraper « Time from last report »)."""
        target = _norm(label)
        prefix_hit: int | None = None
        for i in range(start, len(self.rows)):
            row = self.rows[i]
            if col >= len(row):
                continue
            cell = _norm(row[col])
            if cell == target:
                return i
            if prefix_hit is None and cell.startswith(target):
                prefix_hit = i
        return prefix_hit

    def cell(self, r: int | None, c: int) -> Any:
        if r is None or r >= len(self.rows) or c >= len(self.rows[r]):
            return None
        return self.rows[r][c]

    def right_of(self, label: str, *, offset: int = 1, col: int = 0) -> Any:
        r = self.find(label, col=col)
        return self.cell(r, col + offset)

    def value_after(self, label: str, *, row_offset: int = 0) -> Any:
        """Valeur du libellé situé n'importe où sur une ligne (colonne libre)."""
        target = _norm(label)
        for i, row in enumerate(self.rows):
            for c, v in enumerate(row):
                if _norm(v).startswith(target):
                    return self.cell(i + row_offset, c + 1)
        return None


def _dms(deg: Any, minutes: Any, hemi: Any) -> float | None:
    d, m = _num(deg), _num(minutes)
    if d is None:
        return None
    val = d + (m or 0.0) / 60.0
    h = str(hemi or "").strip().upper()
    if h in ("S", "W"):
        val = -val
    return round(val, 5)


def parse_form(rows: list[list[Any]], *, source_file: str) -> dict[str, Any]:
    g = FormGrid(rows)
    out: dict[str, Any] = {"source_file": source_file}
    out["vessel"] = _text(g.right_of("Vessel name"))
    out["voyage_code"] = _text(g.value_after("Voyage number"))
    out["report_type"] = _text(g.right_of("Type of report"))
    out["eu_mrv"] = _text(g.value_after("Subject to EU-MRV"))
    out["vessel_is"] = _text(g.value_after("Vessel is"))
    out["date_local"] = _text(g.right_of("Date"))
    r_time = g.find("Time")
    out["time_local"] = _text(g.cell(r_time, 1))
    out["tz_label"] = _text(g.cell(r_time, 2))
    out["tz_offset_h"] = _tz_offset(out["tz_label"])
    # Fuseau illisible (« UTC-3:30 », cellule vide…) → pas d'UTC inventé :
    # ``datetime_utc`` reste vide et le rapport est marqué, jamais faussé.
    out["tz_unresolved"] = out["tz_offset_h"] is None
    out["datetime_utc"] = (
        None
        if out["tz_unresolved"]
        else _to_utc(out["date_local"], out["time_local"], out["tz_offset_h"])
    )
    r_lat, r_lon = g.find("Latitude"), g.find("Longitude")
    out["latitude"] = _dms(g.cell(r_lat, 1), g.cell(r_lat, 2), g.cell(r_lat, 3))
    out["longitude"] = _dms(g.cell(r_lon, 1), g.cell(r_lon, 2), g.cell(r_lon, 3))
    out["previous_port"] = _text(g.right_of("Previous port"))
    out["next_port"] = _text(g.right_of("Next Port"))
    out["vessel_condition"] = _text(g.value_after("Vessel condition"))
    out["cargo_quantity_t"] = _num(g.value_after("Cargo quantity"))
    out["draft_fwd_m"] = _num(g.value_after("Draft Fwd"))
    out["draft_aft_m"] = _num(g.value_after("Draft Aft"))
    out["hours_since_last_report"] = _num(g.right_of("Time from last report"))
    out["distance_since_last_report_nm"] = _num(g.right_of("Distance from last report"))
    out["speed_since_last_report_kn"] = _num(g.value_after("Speed from last report"))
    out["hours_since_departure"] = _num(
        g.right_of("Time from departure") or g.right_of("Time from SOSP")
    )
    out["distance_since_departure_nm"] = _num(
        g.right_of("Distance from departure") or g.right_of("Distance from SOSP")
    )
    out["speed_since_sosp_kn"] = _num(g.value_after("Speed from SOSP"))
    out["distance_to_go_nm"] = _num(g.right_of("Distance to go"))
    r_eta = g.find("Announced ETA")
    out["announced_eta"] = {
        "date": _text(g.cell(r_eta, 1)),
        "time": _text(g.cell(r_eta, 2)),
        "tz": _text(g.cell(r_eta, 3)),
    }
    # Engins : ligne du libellé → colonnes 1 (heures) et 2 (conso), + compteurs
    # D / D-1 (Rev 2.1) situés à droite sur la même ligne.
    engines: dict[str, dict[str, Any]] = {}
    r_head = g.find("Running hours", col=1)
    for name in ENGINES:
        r = g.find(name, start=r_head or 0)
        if r is None:
            continue
        row = g.rows[r]
        eng: dict[str, Any] = {
            "running_hours": _num(g.cell(r, 1)),
            "consumption_t": _num(g.cell(r, 2)),
        }
        # Compteurs Rev 2.1 : second libellé identique plus à droite.
        for c in range(3, len(row)):
            if _norm(row[c]).startswith(_norm(name)):
                eng.update(
                    {
                        "counter_hours_d": _num(g.cell(r, c + 1)),
                        "counter_hours_d1": _num(g.cell(r, c + 2)),
                        "counter_litres_d": _num(g.cell(r, c + 3)),
                        "counter_litres_d1": _num(g.cell(r, c + 4)),
                        "counter_hours_departure": _num(g.cell(r, c + 5)),
                        "counter_litres_departure": _num(g.cell(r, c + 6)),
                    }
                )
                break
        engines[name] = eng
    out["engines"] = engines
    r_tot = g.find("Total consumption")
    out["total_consumption_t"] = _num(g.cell(r_tot, 2))
    out["total_consumption_l_per_h"] = _num(g.cell(r_tot, 3))
    out["go_density_t_m3"] = _num(g.value_after("GO Density"))
    out["bunkering_t"] = _num(g.right_of("Bunkering"))
    out["rob_do_t"] = _num(g.right_of("ROB DO"))
    out["rob_urea_t"] = _num(g.right_of("ROB Uree"))
    out["rob_fw_t"] = _num(g.right_of("ROB FW"))
    out["fw_production_t"] = _num(g.right_of("PRODUCTION FW"))
    # Météo par tranche de 4 h.
    r_w = g.find("Weather")
    weather = []
    if r_w is not None:
        for k, slot in enumerate(SLOTS, start=1):
            r = r_w + k
            if _norm(g.cell(r, 0)) != _norm(slot):
                continue
            weather.append(
                {
                    "slot": slot,
                    "tws_kn": _num(g.cell(r, 1)),
                    "awa_deg": _num(g.cell(r, 2)),
                    "aws_kn": _num(g.cell(r, 3)),
                    "sea_state": _num(g.cell(r, 4)),
                    "sea_direction_deg": _num(g.cell(r, 5)),
                    "ship_speed_kn": _num(g.cell(r, 6)),
                }
            )
    out["weather"] = weather
    # Voiles et moteurs par tranche.
    r_s = g.find("Sails")
    sails = []
    if r_s is not None:
        head = [_norm(x) for x in g.rows[r_s]]
        for k, slot in enumerate(SLOTS, start=1):
            r = r_s + k
            if _norm(g.cell(r, 0)) != _norm(slot):
                continue
            entry: dict[str, Any] = {"slot": slot}
            for c, h in enumerate(head[1:], start=1):
                if not h:
                    continue
                v = g.cell(r, c)
                key = h.replace(" ", "_")
                entry[key] = (
                    _on_off(v)
                    if key in ("j0", "fwd_j1", "fwd_ms", "aft_j1", "aft_ms")
                    else (_num(v) if v not in (None, "") else None)
                )
            sails.append(entry)
    out["sails_engines"] = sails
    # Cales : Temperature / Rel. Humidity × Midnight / Midday.
    holds: dict[str, dict[str, Any]] = {}
    r_h = g.find("Midnight", col=1)
    for name in HOLDS:
        r = g.find(name, start=r_h or 0)
        if r is None:
            continue
        holds[name] = {
            "temp_midnight_c": _num(g.cell(r, 1)),
            "rh_midnight_pct": _num(g.cell(r, 2)),
            "temp_midday_c": _num(g.cell(r, 3)),
            "rh_midday_pct": _num(g.cell(r, 4)),
        }
    out["holds"] = holds
    r_c = g.find("Comments")
    out["comments"] = _text(g.cell(r_c + 1, 0)) if r_c is not None else None
    out["form_version"] = _detect_version(rows)
    return out


def _on_off(v: Any) -> bool | None:
    s = str(v or "").strip().upper()
    return True if s == "ON" else False if s == "OFF" else None


def _detect_version(rows: list[list[Any]]) -> str | None:
    for row in rows:
        for v in row:
            s = str(v or "")
            if "Rev" in s and "CFOTE" in " ".join(str(x or "") for x in row):
                return " ".join(str(x or "").strip() for x in row if x).strip()
            if s.startswith("Version"):
                return s.strip()
    return None


def _to_utc(date_local: str | None, time_local: str | None, tz_h: int | None) -> str | None:
    if not date_local:
        return None
    d: datetime | None = None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(date_local[:19], fmt)
            break
        except ValueError:
            continue
    if d is None:
        return None
    hh, mm = 12, 0
    if time_local:
        m = re.match(r"(\d{1,2}):(\d{2})", time_local if ":" in time_local else time_local[11:16])
        if m:
            hh, mm = int(m.group(1)), int(m.group(2))
    local = d.replace(hour=hh, minute=mm, second=0, microsecond=0)
    return (local - timedelta(hours=tz_h or 0)).replace(tzinfo=UTC).isoformat()


def read_workbook_rows(path: Path) -> list[list[Any]]:
    if openpyxl is None:
        raise SystemExit("openpyxl requis : pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Reporting form"] if "Reporting form" in wb.sheetnames else wb.worksheets[0]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()  # mode read_only : le zip reste ouvert sinon (des milliers de fichiers)


def extract(source: Path, out: Path, *, vessel_filter: str | None = None) -> dict[str, int]:
    out.mkdir(parents=True, exist_ok=True)
    files = sorted(
        p
        for p in source.rglob("*.xlsx")
        if not p.name.startswith("~$") and (FILE_RE.match(p.name) or FILE_RE_COMPACT.match(p.name))
    )
    stats = {"files": len(files), "in_scope": 0, "parsed": 0, "failed": 0, "skipped_vessel": 0}
    summary_rows: list[dict[str, Any]] = []
    with (out / "noon_reports.ndjson").open("w", encoding="utf-8") as nd:
        for path in files:
            m = FILE_RE.match(path.name) or FILE_RE_COMPACT.match(path.name)
            assert m is not None
            if vessel_filter and m.group("vessel").upper() != vessel_filter.upper():
                stats["skipped_vessel"] += 1
                continue
            stats["in_scope"] += 1
            try:
                rows = read_workbook_rows(path)
                rec = parse_form(rows, source_file=path.name)
            except Exception as exc:
                stats["failed"] += 1
                print(f"✖ {path.name}: {exc}", file=sys.stderr)
                continue
            rec["file_date"] = m.group("date")
            rec["file_kind"] = m.group("kind").capitalize()
            rec["sha256"] = hashlib.sha256(path.read_bytes()).hexdigest()
            rec["relative_path"] = str(path.relative_to(source))
            nd.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")
            summary_rows.append(
                {
                    "vessel": rec["vessel"],
                    "voyage_code": rec["voyage_code"],
                    "report_type": rec["report_type"] or rec["file_kind"],
                    "file_date": rec["file_date"],
                    "datetime_utc": rec["datetime_utc"],
                    "latitude": rec["latitude"],
                    "longitude": rec["longitude"],
                    "previous_port": rec["previous_port"],
                    "next_port": rec["next_port"],
                    "distance_since_last_report_nm": rec["distance_since_last_report_nm"],
                    "total_consumption_t": rec["total_consumption_t"],
                    "rob_do_t": rec["rob_do_t"],
                    "form_version": rec["form_version"],
                    "tz_unresolved": rec["tz_unresolved"],
                    "source_file": rec["source_file"],
                }
            )
            stats["parsed"] += 1
    with (out / "noon_reports_summary.csv").open("w", encoding="utf-8", newline="") as fh:
        fields = list(summary_rows[0].keys()) if summary_rows else ["source_file"]
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(summary_rows)
    return stats


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--vessel", help="ANEMOS | ARTEMIS | … (défaut : tous)")
    args = parser.parse_args(argv)
    if not args.source.is_dir():
        print(f"✖ dossier introuvable : {args.source}", file=sys.stderr)
        return 2
    stats = extract(args.source, args.out, vessel_filter=args.vessel)
    print(
        f"✔ {stats['parsed']}/{stats['in_scope']} rapports extraits, {stats['failed']} échec(s)"
        + (f", {stats['skipped_vessel']} hors filtre navire" if stats["skipped_vessel"] else "")
    )
    return 0 if not stats["failed"] else 1


if __name__ == "__main__":
    sys.exit(main())
