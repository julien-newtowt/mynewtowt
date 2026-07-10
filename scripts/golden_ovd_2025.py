#!/usr/bin/env python3
"""Golden OVDLA / OVDBR 2025 — critère d'acceptation du LOT 10.

Vérifie que les datasets régénérés depuis les événements importés (dataset
2025) reproduisent les échantillons DNV ``08 - DNV MRV Exports`` :

1. **(a) import** — rejoue ``scripts/import_mrv_2025.py`` (idempotent) sur la
   base cible (``--database-url``) : référentiels + voyages + événements +
   soutages du ``Sample_Dataset_Architecture_Evenementielle_2025.xlsx`` ;
2. **(b) régénération** — reconstruit l'OVDLA/OVDBR ANEMOS (par défaut) via
   ``services.mrv_dataset`` sur la période **d'intersection** réellement
   couverte par les événements importés ;
3. **(c) comparaison LIGNE À LIGNE** aux échantillons, en ne **tolérant que**
   ``Source_System`` (décision Q10 : ``MyTOWT`` vs ``OVDAdmin``) et
   ``Last_Updated`` (horodatage de génération, non données). TOUTE autre
   divergence est LISTÉE, colonne par colonne.

Usage::

    python scripts/golden_ovd_2025.py \\
        --database-url 'postgresql+asyncpg://towt:...@localhost:5432/towt_l1' \\
        --xlsx '.../Sample_Dataset_Architecture_Evenementielle_2025.xlsx' \\
        --samples-dir '.../08 - DNV MRV Exports' \\
        [--vessel ANEMOS] [--skip-import]

Sortie : compte lignes comparées / identiques / divergentes + nature des
divergences par colonne + lignes manquantes/en trop. Code retour 0 (toujours —
le golden est un rapport d'écart, jamais un gate bloquant : les divergences de
reconstruction 2025 sont documentées, cf. rapport du lot).
"""

from __future__ import annotations

import argparse
import asyncio
import os
import secrets
import sys
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import openpyxl

_FLOAT_EPS = 1e-6
# Colonnes tolérées (jamais comptées comme divergence) : Q10 + métadonnée.
_TOLERATED = {"Source_System", "Last_Updated"}


def _norm(col: str, value):
    """Normalise une valeur (échantillon OU régénérée) pour comparaison."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date() if col in ("Date_UTC", "Bunker_Delivery_Date") else value
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip()


def _classify(col: str, a, b) -> str:
    """Classe l'écart d'une colonne : exact | rounding | material | structural.

    - ``exact`` : valeurs égales (ε flottant) ;
    - ``rounding`` : numériques proches (|Δ| ≤ 0,01 ou ≤ 0,1 % relatif) — typiquement
      la précision de stockage (ROB ``Numeric(12,3)`` vs 4 décimales de
      l'échantillon, temps arrondi au 1/1000 h) ;
    - ``structural`` : l'un est absent (None) et l'autre non (ex. 1re ligne dont
      le delta référence une période antérieure absente de l'import) ;
    - ``material`` : écart réel (distance haversine vs distance journalisée ;
      conso reconstruite ≠ conso d'origine)."""
    na, nb = _norm(col, a), _norm(col, b)
    if na is None and nb is None:
        return "exact"
    if (na is None) != (nb is None):
        return "structural"
    if isinstance(na, float) and isinstance(nb, float):
        diff = abs(na - nb)
        if diff <= max(_FLOAT_EPS, abs(nb) * 1e-9):
            return "exact"
        if diff <= 0.01 or diff <= abs(nb) * 0.001:
            return "rounding"
        return "material"
    return "exact" if na == nb else "material"


def _load_sample(path: Path) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in r):
            continue
        rows.append(dict(zip(header, r)))
    return header, rows


def _ovdla_key(row: dict) -> tuple:
    d = _norm("Date_UTC", row.get("Date_UTC"))
    return (str(row.get("Event")), d, str(row.get("Time_UTC")))


def _ovdbr_key(row: dict) -> tuple:
    return (str(row.get("BDN_Number")),)


async def _regenerate(database_url: str, vessel_name: str):
    from sqlalchemy import select
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.models.vessel import Vessel
    from app.services import mrv_dataset as md

    engine = create_async_engine(database_url, echo=False, pool_pre_ping=True)
    sf = async_sessionmaker(engine, expire_on_commit=False, autoflush=False)
    try:
        async with sf() as db:
            vessel = (
                await db.execute(select(Vessel).where(Vessel.name.ilike(vessel_name)))
            ).scalars().first()
            if vessel is None:
                raise SystemExit(f"Navire {vessel_name!r} absent de la base (import ?)")
            # Période d'intersection = amplitude réelle des événements importés.
            from app.models.nav_event import NavEvent

            dts = [
                d
                for d in (
                    await db.execute(
                        select(NavEvent.datetime_utc).where(NavEvent.vessel_id == vessel.id)
                    )
                ).scalars().all()
                if d is not None
            ]
            if not dts:
                raise SystemExit("Aucun événement importé pour ce navire.")
            start = min(dts).replace(tzinfo=UTC) if min(dts).tzinfo is None else min(dts)
            end = max(dts).replace(tzinfo=UTC) if max(dts).tzinfo is None else max(dts)
            ovdla = await md.build_ovdla_rows(db, vessel, start, end)
            ovdbr = await md.build_ovdbr_rows(db, vessel, (start, end))
            return (
                vessel.name,
                (start, end),
                [r.values for r in ovdla if r.included and not r.synthetic],
                [r.values for r in ovdbr if r.included],
            )
    finally:
        await engine.dispose()


def _compare(kind: str, columns, sample_rows, mine_rows, keyfn):
    sample_by = {keyfn(r): r for r in sample_rows}
    mine_by = {keyfn(r): r for r in mine_rows}
    common = [k for k in sample_by if k in mine_by]
    missing = [k for k in sample_by if k not in mine_by]  # dans l'échantillon, pas régénéré
    extra = [k for k in mine_by if k not in sample_by]

    identical = 0          # toutes colonnes exactes
    within_rounding = 0    # aucune divergence matérielle/structurelle (arrondi seul)
    divergent_rows = 0     # ≥1 colonne matérielle ou structurelle
    col_stats: dict[str, dict[str, int]] = {}
    examples: list[str] = []
    for k in common:
        s, m = sample_by[k], mine_by[k]
        classes = []
        row_material = []
        for col in columns:
            if col in _TOLERATED:
                continue
            cls = _classify(col, s.get(col), m.get(col))
            classes.append(cls)
            if cls != "exact":
                col_stats.setdefault(col, {}).setdefault(cls, 0)
                col_stats[col][cls] += 1
            if cls in ("material", "structural"):
                row_material.append(f"{col}[{cls}]: éch={s.get(col)!r}≠rég={m.get(col)!r}")
        if all(c == "exact" for c in classes):
            identical += 1
        elif not row_material:
            within_rounding += 1
        else:
            divergent_rows += 1
            if len(examples) < 10:
                examples.append(f"  [{k}] " + "; ".join(row_material))

    print(f"\n{'='*78}\n{kind}\n{'='*78}")
    print(f"  lignes échantillon : {len(sample_rows)} | régénérées (incluses) : {len(mine_rows)}")
    print(f"  lignes comparées (clé commune) : {len(common)}")
    print(f"    · identiques (toutes colonnes exactes)          : {identical}")
    print(f"    · identiques à l'arrondi près (aucune div. mat.) : {within_rounding}")
    print(f"    · divergentes (≥1 colonne matérielle/structurelle): {divergent_rows}")
    if missing:
        print(f"  lignes de l'échantillon NON régénérées : {len(missing)} → {missing[:6]}")
    if extra:
        print(f"  lignes régénérées ABSENTES de l'échantillon : {len(extra)} → {extra[:6]}")
    if col_stats:
        print("  écarts par colonne (rounding / material / structural) :")
        for col in columns:
            if col in col_stats:
                st = col_stats[col]
                print(f"    - {col:28s} : rounding={st.get('rounding',0)} "
                      f"material={st.get('material',0)} structural={st.get('structural',0)}")
    if examples:
        print("  exemples de divergences matérielles/structurelles :")
        for ex in examples:
            print(ex)
    return {
        "compared": len(common),
        "identical": identical,
        "within_rounding": within_rounding,
        "divergent": divergent_rows,
        "missing": len(missing),
        "extra": len(extra),
    }


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description="Golden OVDLA/OVDBR 2025 (lot 10)")
    parser.add_argument("--database-url", required=True)
    parser.add_argument("--xlsx", required=True, help="Sample_Dataset xlsx")
    parser.add_argument("--samples-dir", required=True, help="Dossier 08 - DNV MRV Exports")
    parser.add_argument("--vessel", default="ANEMOS")
    parser.add_argument("--skip-import", action="store_true", help="Ne pas ré-importer (base déjà peuplée)")
    args = parser.parse_args(argv)

    db_url = args.database_url
    if db_url.startswith("postgresql://"):
        db_url = db_url.replace("postgresql://", "postgresql+asyncpg://", 1)
    os.environ.setdefault("SECRET_KEY", secrets.token_hex(32))
    os.environ.setdefault("DATABASE_URL", db_url)

    from scripts.import_mrv_2025 import load_dataset, run_import

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"ERREUR : xlsx introuvable : {xlsx}", file=sys.stderr)
        return 2
    ds = load_dataset(xlsx)

    if not args.skip_import:
        print(f"(a) import du dataset 2025 (navire {args.vessel}) sur la base cible…")
        import_args = SimpleNamespace(
            database_url=db_url, dry_run=False, vessel=args.vessel.upper(),
            reconcile=False, emit_fixtures=None,
        )
        asyncio.run(run_import(import_args, ds))

    print(f"\n(b) régénération OVDLA/OVDBR {args.vessel}…")
    vessel_name, (start, end), mine_ovdla, mine_ovdbr = asyncio.run(
        _regenerate(db_url, args.vessel)
    )
    print(f"  période d'intersection constatée : {start.date()} → {end.date()} "
          f"({len(mine_ovdla)} lignes OVDLA, {len(mine_ovdbr)} OVDBR régénérées)")

    samples = Path(args.samples_dir)
    ovdla_sample = samples / f"OVDLA 2026-06-15 {vessel_name.title()}.xlsx"
    ovdbr_sample = samples / f"OVDBR 2026-06-15 {vessel_name.title()}.xlsx"

    from app.services.mrv_dataset import OVDBR_COLUMNS, OVDLA_COLUMNS

    print("\n(c) comparaison ligne à ligne (tolérées : Source_System [Q10], Last_Updated)")
    if ovdla_sample.exists():
        _, s_rows = _load_sample(ovdla_sample)
        # L'échantillon peut couvrir au-delà de 2025 : on restreint à l'intersection.
        s_rows = [
            r for r in s_rows
            if str(r.get("Event")) != "Period last event"
            and _in_range(r.get("Date_UTC"), start, end)
        ]
        _compare("OVDLA", OVDLA_COLUMNS, s_rows, mine_ovdla, _ovdla_key)
    else:
        print(f"  ⚠ échantillon OVDLA introuvable : {ovdla_sample}")

    if ovdbr_sample.exists():
        _, s_rows = _load_sample(ovdbr_sample)
        s_rows = [r for r in s_rows if _in_range(r.get("Bunker_Delivery_Date"), start, end)]
        _compare("OVDBR", OVDBR_COLUMNS, s_rows, mine_ovdbr, _ovdbr_key)
    else:
        print(f"  ⚠ échantillon OVDBR introuvable : {ovdbr_sample}")

    return 0


def _in_range(d, start: datetime, end: datetime) -> bool:
    if d is None:
        return False
    dd = d.date() if isinstance(d, datetime) else d
    return start.date() <= dd <= end.date()


if __name__ == "__main__":
    sys.exit(main())
