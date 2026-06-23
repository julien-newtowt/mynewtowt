"""ADM-05 — import en masse d'utilisateurs depuis un classeur Excel.

Onboarding collaborateurs : au lieu de créer les comptes un à un, l'admin
téléverse un classeur (colonnes ``username, email, full_name, role, language,
password``). Chaque ligne est validée indépendamment ; un rapport liste les
comptes créés (avec mot de passe temporaire généré si non fourni) et les lignes
ignorées avec leur motif. Tous les comptes importés sont forcés au changement
de mot de passe à la 1re connexion.
"""

from __future__ import annotations

import io
import re
import secrets

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import hash_password
from app.i18n import SUPPORTED as SUPPORTED_LANGS
from app.models.user import User
from app.permissions import ROLES

# Colonnes du modèle d'import (l'en-tête est insensible à la casse).
IMPORT_COLUMNS: tuple[str, ...] = (
    "username",
    "email",
    "full_name",
    "role",
    "language",
    "password",
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_HEADER_FILL = PatternFill(start_color="0D5966", end_color="0D5966", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


def build_template_xlsx() -> bytes:
    """Classeur modèle (en-têtes seules) pour l'import d'utilisateurs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "USERS"
    for idx, header in enumerate(IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 2)
    return _serialize(wb)


def _serialize(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_users_xlsx(content: bytes) -> list[dict]:
    """Relit le classeur → liste de dicts (clé = nom de colonne connu)."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    known = {c: c for c in IMPORT_COLUMNS}
    col_field: dict[int, str] = {}
    for idx, name in enumerate(header_row):
        if name is None:
            continue
        field = known.get(str(name).strip().lower())
        if field is not None:
            col_field[idx] = field
    out: list[dict] = []
    for row in rows_iter:
        if row is None or not any(v not in (None, "") for v in row):
            continue
        record: dict = {}
        for idx, field in col_field.items():
            if idx < len(row) and row[idx] is not None:
                record[field] = str(row[idx]).strip()
        out.append(record)
    wb.close()
    return out


async def import_users(db: AsyncSession, rows: list[dict]) -> dict:
    """Valide et crée les utilisateurs ligne à ligne. Retourne un rapport.

    Validation : username & email présents, email bien formé, rôle dans
    ``ROLES``, pas de doublon (dans le fichier ou en base). Mot de passe
    temporaire généré si absent ; ``must_change_password=True`` systématique.
    """
    created: list[dict] = []
    skipped: list[dict] = []
    seen_users: set[str] = set()
    seen_emails: set[str] = set()
    for i, row in enumerate(rows, start=2):  # ligne 1 = en-tête
        username = (row.get("username") or "").strip()
        email = (row.get("email") or "").strip().lower()
        role = (row.get("role") or "").strip()
        if not username or not email:
            skipped.append({"line": i, "reason": "username/email manquant"})
            continue
        if not _EMAIL_RE.match(email):
            skipped.append({"line": i, "reason": f"email invalide : {email}"})
            continue
        if role not in ROLES:
            skipped.append({"line": i, "reason": f"rôle invalide : {role or '∅'}"})
            continue
        if username in seen_users or email in seen_emails:
            skipped.append({"line": i, "reason": "doublon dans le fichier"})
            continue
        existing = await db.scalar(
            select(User.id).where((User.username == username) | (User.email == email))
        )
        if existing is not None:
            skipped.append({"line": i, "reason": "utilisateur déjà existant"})
            continue
        provided_pwd = (row.get("password") or "").strip()
        temp_pwd = provided_pwd if len(provided_pwd) >= 12 else secrets.token_urlsafe(12)
        lang = (row.get("language") or "fr").strip().lower()
        if lang not in SUPPORTED_LANGS:
            lang = "fr"
        u = User(
            username=username,
            email=email,
            full_name=(row.get("full_name") or "").strip() or None,
            hashed_password=hash_password(temp_pwd),
            role=role,
            language=lang,
            must_change_password=True,
        )
        db.add(u)
        await db.flush()
        seen_users.add(username)
        seen_emails.add(email)
        created.append(
            {
                "username": username,
                "email": email,
                "role": role,
                # Mot de passe temporaire à transmettre à l'utilisateur (changé
                # à la 1re connexion). Affiché une seule fois dans le rapport.
                "temp_password": None if provided_pwd else temp_pwd,
            }
        )
    return {"created": created, "skipped": skipped}
