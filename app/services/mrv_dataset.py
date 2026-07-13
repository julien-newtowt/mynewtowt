"""Sorties réglementaires OVDLA / OVDBR (LOT 10) — l'UNIQUE générateur.

Produit les deux datasets déposés chez DNV, qui remplacent l'export CSV
18 colonnes (déprécié ici, retiré au lot 14) :

- **OVDLA** (*Voyage Log Abstract*) : 1 ligne par événement de navigation
  **validé** (Departure / Arrival / Begin|End Anchoring — **jamais** les Noon,
  Q10), en **DELTAS depuis l'événement précédent du navire** (tous types
  confondus pour le calcul : les deltas de conso/distance/temps AGRÈGENT les
  intervalles — noon intermédiaires inclus — depuis le dernier événement OVDLA).
  Une ligne synthétique « Period last event » clôt la période de reporting
  quand un voyage est ouvert au ``period_end`` (Q10).
- **OVDBR** (*Bunker Report*) : 1 ligne par soutage **validé Master**.

Structure/format des colonnes REPRODUITS des échantillons 2025 DNV (dissection
au rapport §OVDLA/OVDBR) : minutes DMS **entières**, ``Time_UTC`` en ``HH:MM``,
``Date_UTC`` date seule, consommations en tonnes. ``Source_System = "MyTOWT"``
(Q10, l'émetteur du format est MyTOWT — « OVDAdmin » observé venait de l'ancien
outil). Calculs délégués à ``inter_event_compute`` (deltas de compteurs, ROB
chaîné) et ``mrv_compute.decimal_to_dms`` (positions).

**Portes de génération** (la qualité est une PORTE, pas un rapport a posteriori) :

- OVDLA : seuls les événements ``valide`` entrent ; un événement dont un rapport
  lié est ``under_conformity`` (``report_generation.report_quality_status``) est
  EXCLU **et déclenche une alerte** admin (pattern lot 8) ; idem si son entrée
  gelée porte déjà ``under_conformity`` ;
- OVDBR : seuls les soutages ``valide_master`` entrent ; entrée gelée
  ``under_conformity`` ⇒ exclusion + alerte.

Les lignes sont **gelées** dans ``mrv_log_abstract_entries`` /
``mrv_bunkering_entries`` par ``snapshot_entries`` (upsert idempotent,
``verification_status`` conservé). Exports ``export_xlsx`` (openpyxl, mise en
forme sobre type échantillon) et ``export_csv``.
"""

from __future__ import annotations

import csv
import io
import itertools
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bunker import BunkerOperation
from app.models.env_report import EnvReportEventLink
from app.models.leg import Leg
from app.models.mrv_dataset import (
    SOURCE_SYSTEM_DEFAULT,
    MrvBunkeringEntry,
    MrvLogAbstractEntry,
)
from app.models.nav_event import (
    NavEvent,
    PortCallEvent,
)
from app.models.port import Port
from app.models.vessel import Vessel
from app.models.vessel_env import VesselEngine
from app.services import inter_event_compute as iec
from app.services import notifications
from app.services.mrv_compute import decimal_to_dms
from app.utils.csv_safe import sanitize_cell

# ════════════════════════════════════════════════════════════ Colonnes / labels

# En-têtes OVDLA — ordre et intitulés EXACTS des échantillons DNV 2025.
OVDLA_COLUMNS: tuple[str, ...] = (
    "IMO",
    "Date_UTC",
    "Time_UTC",
    "Event",
    "Time_Since_Previous_Report",
    "Distance",
    "Latitude_North_South",
    "Latitude_Degree",
    "Latitude_Minutes",
    "Longitude_East_West",
    "Longitude_Degree",
    "Longitude_Minutes",
    "Voyage_From",
    "Voyage_To",
    "Cargo_Mt",
    "ME_Consumption_MDO",
    "AE_Consumption_MDO",
    "MDO_ROB",
    "Source_System",
    "Last_Updated",
)

# En-têtes OVDBR — ordre et intitulés EXACTS des échantillons DNV 2025.
OVDBR_COLUMNS: tuple[str, ...] = (
    "IMO",
    "BDN_Number",
    "Bunker_Delivery_Date",
    "Bunker_Delivery_Time",
    "Bunker_Port",
    "Fuel_Type",
    "Mass",
    "Source_System",
    "Last_Updated",
)

# Type d'événement (interne) → libellé « Event » de l'OVDLA. Les Noon ne
# produisent PAS de ligne (Q10) : ils n'apparaissent pas dans ce mapping.
OVDLA_EVENT_LABELS: dict[str, str] = {
    "departure": "Departure",
    "arrival": "Arrival",
    "anchoring_begin": "Begin Anchoring/Drifting",
    "anchoring_end": "End Anchoring/Drifting",
}
# Types d'événement produisant une ligne OVDLA (non-noon, validés).
OVDLA_ROW_EVENT_TYPES: tuple[str, ...] = tuple(OVDLA_EVENT_LABELS)
# Libellé de la ligne synthétique de clôture de période (Q10).
PERIOD_LAST_EVENT_LABEL = "Period last event"

_HOUR_QUANT = Decimal("0.001")


# ════════════════════════════════════════════════════════════ Résultat typé


@dataclass
class DatasetRow:
    """Une ligne candidate d'un dataset (OVDLA/OVDBR) + son verdict de porte.

    ``values`` porte les colonnes gelées (en-têtes exacts → valeurs natives).
    ``included`` = passe les portes (sinon ``exclusion_reason`` motive le rejet,
    affiché à l'aperçu). ``synthetic`` marque la ligne « Period last event »
    (pas d'événement rattaché → non persistée, présente à l'export seul).
    """

    kind: str  # "ovdla" | "ovdbr"
    values: dict[str, Any]
    event_id: int | None = None
    bunker_id: int | None = None
    verification_status: str = "conform"
    included: bool = True
    exclusion_reason: str | None = None
    synthetic: bool = False


# ════════════════════════════════════════════════════════════ Helpers


def _naive_utc(dt: datetime | None) -> datetime | None:
    """Normalise en naïf-UTC (compare uniformément PG aware / SQLite naïf)."""
    if dt is None:
        return None
    return dt.astimezone(UTC).replace(tzinfo=None) if dt.tzinfo is not None else dt


def _in_period(dt: datetime | None, start: datetime | None, end: datetime | None) -> bool:
    d = _naive_utc(dt)
    if d is None:
        return False
    if start is not None and d < _naive_utc(start):
        return False
    return not (end is not None and d > _naive_utc(end))


def _dms_row(lat: Decimal | None, lon: Decimal | None) -> dict[str, Any]:
    """Position décimale → colonnes DMS de l'OVDLA (minutes ENTIÈRES, cf. échantillons)."""
    out: dict[str, Any] = {
        "Latitude_North_South": None,
        "Latitude_Degree": None,
        "Latitude_Minutes": None,
        "Longitude_East_West": None,
        "Longitude_Degree": None,
        "Longitude_Minutes": None,
    }
    if lat is not None:
        deg, minutes, hemi = decimal_to_dms(float(lat), is_lat=True)
        out["Latitude_North_South"] = hemi
        out["Latitude_Degree"] = deg
        out["Latitude_Minutes"] = int(minutes.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if lon is not None:
        deg, minutes, hemi = decimal_to_dms(float(lon), is_lat=False)
        out["Longitude_East_West"] = hemi
        out["Longitude_Degree"] = deg
        out["Longitude_Minutes"] = int(minutes.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return out


def _hhmm(dt: datetime | None) -> str | None:
    d = _naive_utc(dt)
    return None if d is None else d.strftime("%H:%M")


def _jsonable(value: Any) -> Any:
    """Valeur JSON-safe pour le payload gelé (Decimal→str, date/datetime→ISO)."""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _freeze(values: dict[str, Any]) -> dict[str, Any]:
    return {k: _jsonable(v) for k, v in values.items()}


# ════════════════════════════════════════════════════════════ Chargement chaîne


async def _vessel_chain(db: AsyncSession, vessel_id: int) -> list[NavEvent]:
    """Chaîne complète du navire (tous legs, brouillons exclus), triée par UTC.

    Inclut les événements ANTÉRIEURS à la période demandée : les deltas de la
    1re ligne d'une période se calculent depuis l'événement précédent réel
    (comme l'échantillon dont la 1re ligne référence la période antérieure)."""
    rows = await db.execute(
        select(NavEvent)
        .where(
            NavEvent.vessel_id == vessel_id,
            NavEvent.status.in_(iec.FINALIZED_STATUSES),
            NavEvent.datetime_utc.isnot(None),
        )
        .order_by(NavEvent.datetime_utc.asc(), NavEvent.id.asc())
    )
    return list(rows.scalars().all())


async def _load_engines(db: AsyncSession, vessel_id: int) -> dict[int, VesselEngine]:
    rows = await db.execute(select(VesselEngine).where(VesselEngine.vessel_id == vessel_id))
    return {e.id: e for e in rows.scalars().all()}


async def _leg_port_map(
    db: AsyncSession, leg_ids: set[int]
) -> dict[int, tuple[str | None, str | None]]:
    """{leg_id: (locode_départ, locode_arrivée)} — Voyage_From / Voyage_To."""
    if not leg_ids:
        return {}
    legs = list((await db.execute(select(Leg).where(Leg.id.in_(leg_ids)))).scalars().all())
    port_ids = {p for leg in legs for p in (leg.departure_port_id, leg.arrival_port_id) if p}
    ports: dict[int, Port] = {}
    if port_ids:
        ports = {
            p.id: p
            for p in (await db.execute(select(Port).where(Port.id.in_(port_ids)))).scalars().all()
        }
    out: dict[int, tuple[str | None, str | None]] = {}
    for leg in legs:
        dep = ports.get(leg.departure_port_id) if leg.departure_port_id else None
        arr = ports.get(leg.arrival_port_id) if leg.arrival_port_id else None
        out[leg.id] = (getattr(dep, "locode", None), getattr(arr, "locode", None))
    return out


async def _under_conformity_event_ids(db: AsyncSession, event_ids: set[int]) -> set[int]:
    """Sous-ensemble des ``event_ids`` liés à un rapport ``under_conformity``.

    Un événement lié (``env_report_event_links``) à un rapport dont le pire
    statut qualité est ``under_conformity`` est une PORTE fermée (lot 10)."""
    if not event_ids:
        return set()
    from app.services.report_generation import report_quality_status

    links = (
        await db.execute(
            select(EnvReportEventLink.event_id, EnvReportEventLink.report_id).where(
                EnvReportEventLink.event_id.in_(event_ids)
            )
        )
    ).all()
    blocked: set[int] = set()
    # report_id → statut (mémoïsé : plusieurs événements peuvent partager un rapport).
    status_cache: dict[int, str | None] = {}
    for ev_id, report_id in links:
        if report_id not in status_cache:
            status_cache[report_id] = await report_quality_status(db, report_id)
        if status_cache[report_id] == "under_conformity":
            blocked.add(ev_id)
    return blocked


def _rob_by_event(
    chain: list[NavEvent], engines: dict[int, VesselEngine], density: Decimal
) -> dict[int, Decimal | None]:
    """ROB par événement : déclaré (PortCall) ; sinon chaîné PAR LEG (mouillages).

    Le ROB déclaré des escales est la source de référence (R14-v2) et reproduit
    l'échantillon à l'identique. Le ROB des mouillages n'est pas déclaré : il est
    chaîné DANS SON LEG (``compute_rob_chain`` réancre sur le Departure) avec la
    conso réelle par deltas de compteurs, jamais à travers la remise à zéro des
    compteurs d'un changement de voyage."""
    out: dict[int, Decimal | None] = {}
    by_leg: dict[int, list[NavEvent]] = {}
    for ev in chain:
        by_leg.setdefault(ev.leg_id, []).append(ev)
    for leg_events in by_leg.values():
        intervals = [
            iec.compute_interval(prev, cur, engines, density)
            for prev, cur in itertools.pairwise(leg_events)
        ]
        for point in iec.compute_rob_chain(leg_events, intervals):
            out[point.event_id] = point.rob_calculated_t
    # Le déclaré prime pour les escales (source R14-v2, = échantillon).
    for ev in chain:
        if isinstance(ev, PortCallEvent) and ev.rob_t is not None:
            out[ev.id] = Decimal(ev.rob_t)
    return out


# ════════════════════════════════════════════════════════════ Agrégat deltas


@dataclass
class _Window:
    """Agrégat des intervalles depuis la dernière ligne OVDLA émise."""

    distance_nm: Decimal = Decimal("0")
    me_t: Decimal = Decimal("0")
    ae_t: Decimal = Decimal("0")
    has_anomaly: bool = False

    def add(self, interval: iec.IntervalResult) -> None:
        if interval.distance_nm is not None:
            self.distance_nm += interval.distance_nm
        me = interval.group_conso_t.get("ME")
        ae = interval.group_conso_t.get("AE")
        if me is not None:
            self.me_t += me
        if ae is not None:
            self.ae_t += ae
        if interval.counter_anomaly or me is None or ae is None:
            self.has_anomaly = True

    def reset(self) -> None:
        self.distance_nm = Decimal("0")
        self.me_t = Decimal("0")
        self.ae_t = Decimal("0")
        self.has_anomaly = False


# ════════════════════════════════════════════════════════════ OVDLA


async def build_ovdla_rows(
    db: AsyncSession,
    vessel: Vessel,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
    *,
    alert: bool = False,
) -> list[DatasetRow]:
    """Lignes OVDLA du navire sur la période (DELTAS depuis l'événement précédent).

    1 ligne par événement non-Noon **validé** de la période, avec conso ME/AE,
    distance et temps AGRÉGÉS sur les intervalles (noon inclus) depuis la ligne
    OVDLA précédente. ``period_end`` fourni + voyage ouvert ⇒ ligne synthétique
    « Period last event ». Renvoie AUSSI les lignes exclues (motif renseigné)
    pour l'aperçu ; ``alert=True`` notifie l'admin des exclusions
    ``under_conformity`` (pattern lot 8)."""
    chain = await _vessel_chain(db, vessel.id)
    if not chain:
        return []
    engines = await _load_engines(db, vessel.id)
    density = await iec.resolve_density(db, vessel.id)
    port_map = await _leg_port_map(db, {ev.leg_id for ev in chain})
    rob_map = _rob_by_event(chain, engines, density)
    # Cargo « deadweight carried » du voyage = celui du Departure (calculé par
    # ``compute_cargo_mrv`` — repli sur ``cargo_mrv_t`` saisi tant que les
    # hydrostatiques manquent, Q11) ; reporté sur toutes les lignes du leg
    # (mouillages inclus, comme l'échantillon garde le cargo constant/voyage).
    from app.models.vessel_env import VesselHydrostatics as _VH

    hydro = list((await db.execute(select(_VH).where(_VH.vessel_id == vessel.id))).scalars().all())
    leg_cargo: dict[int, Decimal | None] = {}
    for ev in chain:
        if ev.event_type == "departure" and isinstance(ev, PortCallEvent):
            leg_cargo[ev.leg_id] = iec.compute_cargo_mrv(ev, vessel, hydro).cargo_mrv_t
    row_event_ids = {
        ev.id
        for ev in chain
        if ev.event_type in OVDLA_ROW_EVENT_TYPES
        and ev.status == "valide"
        and _in_period(ev.datetime_utc, period_start, period_end)
    }
    blocked = await _under_conformity_event_ids(db, row_event_ids)
    frozen = await _frozen_status_map(db)

    imo = vessel.imo_number or ""
    now = datetime.now(UTC)
    intervals = [
        iec.compute_interval(prev, cur, engines, density) for prev, cur in itertools.pairwise(chain)
    ]

    rows: list[DatasetRow] = []
    window = _Window()
    last_row_dt: datetime | None = None
    last_seen: NavEvent | None = None  # dernier événement (tous types) vu

    for idx, ev in enumerate(chain):
        # Au-delà de ``period_end`` : événements d'une période ultérieure — on
        # stoppe (chaîne triée). Ne pas les agréger ni avancer le point de
        # contrôle : la fenêtre en cours alimente la ligne « Period last event ».
        if period_end is not None and _naive_utc(ev.datetime_utc) > _naive_utc(period_end):
            break
        if idx > 0:
            window.add(intervals[idx - 1])
        last_seen = ev
        if ev.event_type not in OVDLA_ROW_EVENT_TYPES:
            continue  # Noon : agrège mais ne produit pas de ligne (Q10)
        in_period = _in_period(ev.datetime_utc, period_start, period_end)
        # Porte : événement non validé / under_conformity ⇒ exclu (mais reset).
        excluded_reason: str | None = None
        vstatus = frozen.get(ev.id, "conform")
        if ev.status != "valide":
            excluded_reason = "événement non validé"
        elif ev.id in blocked or vstatus == "under_conformity":
            excluded_reason = "rapport lié en non-conformité (under_conformity)"
            vstatus = "under_conformity"

        if in_period:
            values = _ovdla_values(ev, imo, window, last_row_dt, port_map, rob_map, leg_cargo, now)
            row = DatasetRow(
                kind="ovdla",
                values=values,
                event_id=ev.id,
                verification_status=vstatus,
                included=excluded_reason is None,
                exclusion_reason=excluded_reason,
            )
            rows.append(row)
            if alert and excluded_reason and vstatus == "under_conformity":
                await _alert_exclusion(db, "OVDLA", vessel, str(ev.id), excluded_reason)

        # Reset de la fenêtre à chaque événement OVDLA (émis OU exclu : c'est un
        # point de contrôle structurel du dataset — chaque ligne porte le delta
        # depuis le point OVDLA immédiatement précédent). Le point de référence
        # temporel avance donc sur chaque événement OVDLA rencontré.
        window.reset()
        last_row_dt = _naive_utc(ev.datetime_utc)

    # ── Ligne synthétique « Period last event » (Q10) : voyage ouvert au bord
    # de période (des événements — noon — tombent après la dernière ligne OVDLA
    # mais dans la période, sans qu'aucun événement OVDLA ne coïncide avec
    # ``period_end``).
    if period_end is not None and last_seen is not None:
        tail = [
            ev
            for ev in chain
            if ev.event_type not in OVDLA_ROW_EVENT_TYPES
            and _in_period(ev.datetime_utc, last_row_dt, period_end)
        ]
        last_row_is_at_end = any(
            r.event_id is not None and _naive_utc(_row_dt(r)) == _naive_utc(period_end)
            for r in rows
        )
        if tail and not last_row_is_at_end:
            marker = tail[-1]
            values = _ovdla_values(
                marker,
                imo,
                window,
                last_row_dt,
                port_map,
                rob_map,
                leg_cargo,
                now,
                event_label=PERIOD_LAST_EVENT_LABEL,
                override_dt=period_end,
            )
            rows.append(
                DatasetRow(
                    kind="ovdla",
                    values=values,
                    event_id=None,
                    synthetic=True,
                    included=True,
                )
            )

    return rows


def _row_dt(row: DatasetRow) -> datetime | None:
    d = row.values.get("Date_UTC")
    t = row.values.get("Time_UTC")
    if not isinstance(d, date):
        return None
    hh, mm = (int(x) for x in (t or "00:00").split(":"))
    return datetime(d.year, d.month, d.day, hh, mm)


def _ovdla_values(
    ev: NavEvent,
    imo: str,
    window: _Window,
    last_row_dt: datetime | None,
    port_map: dict[int, tuple[str | None, str | None]],
    rob_map: dict[int, Decimal | None],
    leg_cargo: dict[int, Decimal | None],
    now: datetime,
    *,
    event_label: str | None = None,
    override_dt: datetime | None = None,
) -> dict[str, Any]:
    ev_dt = _naive_utc(override_dt if override_dt is not None else ev.datetime_utc)
    time_since: Decimal | None = None
    if last_row_dt is not None and ev_dt is not None:
        time_since = (
            Decimal(str((ev_dt - last_row_dt).total_seconds())) / Decimal("3600")
        ).quantize(_HOUR_QUANT)
    frm, to = port_map.get(ev.leg_id, (None, None))
    # Cargo propre à l'événement (Departure/Arrival) sinon celui du voyage
    # (mouillages / Period last event) — cargo constant par voyage (échantillon).
    cargo = ev.cargo_mrv_t if ev.cargo_mrv_t is not None else leg_cargo.get(ev.leg_id)
    values = {
        "IMO": imo,
        "Date_UTC": ev_dt.date() if ev_dt is not None else None,
        "Time_UTC": ev_dt.strftime("%H:%M") if ev_dt is not None else None,
        "Event": event_label or OVDLA_EVENT_LABELS.get(ev.event_type, ev.event_type),
        "Time_Since_Previous_Report": time_since,
        "Distance": window.distance_nm.quantize(Decimal("0.001")),
        **_dms_row(ev.lat_decimal, ev.lon_decimal),
        "Voyage_From": frm,
        "Voyage_To": to,
        "Cargo_Mt": (Decimal(cargo) if cargo is not None else None),
        "ME_Consumption_MDO": window.me_t.quantize(Decimal("0.00001")),
        "AE_Consumption_MDO": window.ae_t.quantize(Decimal("0.00001")),
        "MDO_ROB": rob_map.get(ev.id),
        "Source_System": SOURCE_SYSTEM_DEFAULT,
        "Last_Updated": now.replace(microsecond=0, tzinfo=None),
    }
    return values


# ════════════════════════════════════════════════════════════ OVDBR


async def build_ovdbr_rows(
    db: AsyncSession,
    vessel: Vessel,
    period: tuple[datetime | None, datetime | None] | None = None,
    *,
    alert: bool = False,
) -> list[DatasetRow]:
    """Lignes OVDBR du navire : 1 par soutage **validé Master** (période optionnelle).

    ``period`` = ``(start, end)`` filtre sur ``delivery_datetime_utc`` (None =
    pas de borne). Porte : ``status == 'valide_master'`` ; un soutage dont
    l'entrée gelée est ``under_conformity`` est exclu (+ alerte si ``alert``)."""
    start, end = period or (None, None)
    bunkers = list(
        (
            await db.execute(
                select(BunkerOperation)
                .where(BunkerOperation.vessel_id == vessel.id)
                .order_by(BunkerOperation.delivery_datetime_utc.asc())
            )
        )
        .scalars()
        .all()
    )
    frozen = await _frozen_bunker_status_map(db)
    imo = vessel.imo_number or ""
    now = datetime.now(UTC).replace(microsecond=0, tzinfo=None)

    rows: list[DatasetRow] = []
    for b in bunkers:
        if not _in_period(b.delivery_datetime_utc, start, end):
            continue
        vstatus = frozen.get(b.id, "conform")
        reason: str | None = None
        if b.status != "valide_master":
            reason = "soutage non validé Master"
        elif vstatus == "under_conformity":
            reason = "entrée en non-conformité (under_conformity)"
        dt = _naive_utc(b.delivery_datetime_utc)
        values = {
            "IMO": imo,
            "BDN_Number": b.bdn_number,
            "Bunker_Delivery_Date": dt.date() if dt is not None else None,
            "Bunker_Delivery_Time": _hhmm(b.delivery_datetime_utc),
            "Bunker_Port": b.port_locode,
            "Fuel_Type": b.fuel_type,
            "Mass": (Decimal(b.mass_t) if b.mass_t is not None else None),
            "Source_System": SOURCE_SYSTEM_DEFAULT,
            "Last_Updated": now,
        }
        rows.append(
            DatasetRow(
                kind="ovdbr",
                values=values,
                bunker_id=b.id,
                verification_status=vstatus,
                included=reason is None,
                exclusion_reason=reason,
            )
        )
        if alert and reason and vstatus == "under_conformity":
            await _alert_exclusion(db, "OVDBR", vessel, b.bdn_number, reason)
    return rows


# ════════════════════════════════════════════════════════════ Statuts gelés


async def _frozen_status_map(db: AsyncSession) -> dict[int, str]:
    rows = (
        await db.execute(
            select(MrvLogAbstractEntry.event_id, MrvLogAbstractEntry.verification_status)
        )
    ).all()
    return dict(rows)


async def _frozen_bunker_status_map(db: AsyncSession) -> dict[int, str]:
    rows = (
        await db.execute(select(MrvBunkeringEntry.bunker_id, MrvBunkeringEntry.verification_status))
    ).all()
    return dict(rows)


# ════════════════════════════════════════════════════════════ Alerte (lot 8)


async def _alert_exclusion(
    db: AsyncSession, dataset: str, vessel: Vessel, subject: str, reason: str
) -> None:
    """Notifie l'administrateur d'une exclusion de consolidation (dédup par lien)."""
    link = f"/mrv/datasets?vessel_id={vessel.id}"
    title = f"Consolidation {dataset} — exclusion ({vessel.name})"
    detail = f"{dataset} · {subject} : {reason}"[:480]
    from app.models.notification import Notification

    exists = (
        await db.execute(
            select(Notification.id)
            .where(
                Notification.link == link,
                Notification.target_role == "administrateur",
                Notification.detail == detail,
                Notification.is_archived.is_(False),
            )
            .limit(1)
        )
    ).first()
    if exists is not None:
        return
    await notifications.create(
        db,
        type="info",
        title=title,
        detail=detail,
        link=link,
        target_role="administrateur",
    )


# ════════════════════════════════════════════════════════════ Snapshot (gel)


async def snapshot_entries(db: AsyncSession, rows: list[DatasetRow]) -> dict[str, int]:
    """Upsert des payloads GELÉS des lignes INCLUSES dans les 2 tables.

    Idempotent (event_id / bunker_id UNIQUE). Le ``verification_status`` d'une
    entrée existante est CONSERVÉ (une vérification manuelle n'est jamais écrasée
    par une régénération) ; seul le payload gelé et ``source_system`` sont
    rafraîchis. Les lignes synthétiques (Period last event) ne sont pas
    persistées (aucun événement à rattacher). Renvoie le compte créé/mis à jour."""
    created = updated = 0
    for row in rows:
        if not row.included or row.synthetic:
            continue
        payload = _freeze(row.values)
        if row.kind == "ovdla" and row.event_id is not None:
            existing = (
                await db.execute(
                    select(MrvLogAbstractEntry).where(MrvLogAbstractEntry.event_id == row.event_id)
                )
            ).scalar_one_or_none()
            if existing is not None:
                existing.payload = payload
                existing.source_system = SOURCE_SYSTEM_DEFAULT
                updated += 1
            else:
                db.add(
                    MrvLogAbstractEntry(
                        event_id=row.event_id,
                        source_system=SOURCE_SYSTEM_DEFAULT,
                        verification_status=row.verification_status,
                        payload=payload,
                    )
                )
                created += 1
        elif row.kind == "ovdbr" and row.bunker_id is not None:
            existing_b = (
                await db.execute(
                    select(MrvBunkeringEntry).where(MrvBunkeringEntry.bunker_id == row.bunker_id)
                )
            ).scalar_one_or_none()
            if existing_b is not None:
                existing_b.payload = payload
                existing_b.source_system = SOURCE_SYSTEM_DEFAULT
                updated += 1
            else:
                db.add(
                    MrvBunkeringEntry(
                        bunker_id=row.bunker_id,
                        source_system=SOURCE_SYSTEM_DEFAULT,
                        verification_status=row.verification_status,
                        payload=payload,
                    )
                )
                created += 1
    await db.flush()
    return {"created": created, "updated": updated}


# ════════════════════════════════════════════════════════════ Exports


def _columns_for(kind: str) -> tuple[str, ...]:
    return OVDLA_COLUMNS if kind == "ovdla" else OVDBR_COLUMNS


def _cell(value: Any) -> Any:
    """Valeur cellule xlsx : Decimal→float, date/datetime tels quels, None→''.

    Anti-injection de formule : une cellule texte (BDN, port, type de
    carburant) commençant par ``= + - @`` serait interprétée comme une formule
    à l'ouverture du fichier déposé chez DNV — ``sanitize_cell`` la neutralise
    (les nombres/dates, non-``str``, passent intacts). Cf. ``app/utils/csv_safe``.
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return sanitize_cell(value)


def _csv_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    # Anti-injection de formule (cf. _cell) sur la valeur texte sérialisée.
    return sanitize_cell(str(value))


def export_xlsx(rows: list[DatasetRow], *, kind: str | None = None) -> bytes:
    """Sérialise les lignes INCLUSES en xlsx (openpyxl, mise en forme sobre).

    ``kind`` déduit des lignes si omis. Formats de date/heure calqués sur les
    échantillons DNV (``yyyy-mm-dd`` / ``yyyy-mm-dd hh:mm``). En-tête en gras."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    resolved = kind or (rows[0].kind if rows else "ovdla")
    columns = _columns_for(resolved)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        if not row.included:
            continue
        line = [_cell(row.values.get(c)) for c in columns]
        ws.append(line)
        r = ws.max_row
        for i, col in enumerate(columns, start=1):
            val = row.values.get(col)
            if col in ("Date_UTC", "Bunker_Delivery_Date") and isinstance(val, date):
                ws.cell(row=r, column=i).number_format = "yyyy-mm-dd"
            elif col == "Last_Updated" and isinstance(val, datetime):
                ws.cell(row=r, column=i).number_format = "yyyy-mm-dd hh:mm"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(rows: list[DatasetRow], *, kind: str | None = None) -> str:
    """Sérialise les lignes INCLUSES en CSV (séparateur virgule, en-têtes exacts)."""
    resolved = kind or (rows[0].kind if rows else "ovdla")
    columns = _columns_for(resolved)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(list(columns))
    for row in rows:
        if not row.included:
            continue
        writer.writerow([_csv_cell(row.values.get(c)) for c in columns])
    return buf.getvalue()
