"""CARGO-09 — import / export Excel de la packing list (saisie de masse).

Un voyage = des dizaines de batches : le canal Excel permet à l'agent cargo
et à l'expéditeur de saisir/exporter en masse. Trois usages :

- **export PL** : un classeur reprenant les batches d'une packing list, avec
  colonnes de contexte (voyage / POL / POD / navire) en lecture seule et
  colonnes éditables ;
- **template** : le même classeur, en-têtes seules, à remplir hors-ligne ;
- **import** : relit un classeur et reconstruit les batches (remplacement).

Le mapping colonne→champ ne porte que sur les champs **éditables** du
``PackingListBatch`` V3 ; les colonnes de contexte sont ignorées à l'import
(elles dérivent du leg). Les valeurs sont typées via ``coerce_batch_form`` —
une cellule vide laisse jouer le défaut de colonne (pas d'écrasement).
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.packing_list import coerce_batch_form
from app.utils.csv_safe import sanitize_cell

# Colonnes de contexte (lecture seule à l'export, ignorées à l'import).
CONTEXT_HEADERS: tuple[str, ...] = (
    "BATCH_NUMBER",
    "VOYAGE_ID",
    "VESSEL",
    "POL_CODE",
    "POD_CODE",
    "BL_NUMBER",
)

# Colonnes éditables : en-tête Excel → champ ``PackingListBatch`` (tous dans
# AUDITABLE_FIELDS, donc typés par ``coerce_batch_form``).
EDITABLE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("PALLET_FORMAT", "pallet_format"),
    ("PALLET_COUNT", "pallet_count"),
    ("HS_CODE", "hs_code"),
    ("TYPE_OF_GOODS", "type_of_goods"),
    ("DESCRIPTION_OF_GOODS", "description_of_goods"),
    ("CASES_QUANTITY", "cases_quantity"),
    ("UNITS_PER_CASE", "units_per_case"),
    ("CARGO_VALUE_USD", "cargo_value_usd"),
    ("WEIGHT_KG", "weight_kg"),
    ("LENGTH_CM", "length_cm"),
    ("WIDTH_CM", "width_cm"),
    ("HEIGHT_CM", "height_cm"),
    ("CUBAGE_M3", "cubage_m3"),
    ("HAZARDOUS", "hazardous"),
    ("IMDG_CLASS", "imdg_class"),
    ("UN_NUMBER", "un_number"),
    ("STACKABLE", "stackable"),
    ("MARKS_AND_NUMBERS", "marks_and_numbers"),
    ("SHIPPER_NAME", "shipper_name"),
    ("SHIPPER_ADDRESS", "shipper_address"),
    ("SHIPPER_CITY", "shipper_city"),
    ("SHIPPER_COUNTRY", "shipper_country"),
    ("CONSIGNEE_NAME", "consignee_name"),
    ("CONSIGNEE_ADDRESS", "consignee_address"),
    ("CONSIGNEE_CITY", "consignee_city"),
    ("CONSIGNEE_COUNTRY", "consignee_country"),
    ("NOTIFY_NAME", "notify_name"),
    ("NOTIFY_ADDRESS", "notify_address"),
)

ALL_HEADERS: tuple[str, ...] = CONTEXT_HEADERS + tuple(h for h, _ in EDITABLE_COLUMNS)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FILL = PatternFill(start_color="0D5966", end_color="0D5966", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOOL_FIELDS = {"hazardous", "stackable"}


def _new_sheet() -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = "PACKING_LIST"
    for col_idx, header in enumerate(ALL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) + 2)
    return wb, ws


def batch_context(batch, *, voyage_id, vessel, pol_code, pod_code) -> dict:
    """Construit le dict de colonnes de contexte d'un batch (export voyage/PL)."""
    return {
        "BATCH_NUMBER": batch.batch_number,
        "VOYAGE_ID": voyage_id,
        "VESSEL": vessel,
        "POL_CODE": pol_code,
        "POD_CODE": pod_code,
        "BL_NUMBER": batch.bl_number,
    }


def _row_values(batch, ctx: dict) -> list:
    values: list = [ctx.get(h) for h in CONTEXT_HEADERS]
    for _header, field in EDITABLE_COLUMNS:
        raw = getattr(batch, field, None)
        if field in _BOOL_FIELDS:
            raw = 1 if raw else 0
        values.append(raw)
    # Anti-injection de formule : une cellule texte saisie par l'utilisateur
    # commençant par = + - @ serait interprétée comme formule à l'ouverture.
    return [sanitize_cell(v) for v in values]


def _serialize(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_xlsx() -> bytes:
    """Classeur vide (en-têtes seules) à remplir hors-ligne."""
    wb, _ws = _new_sheet()
    return _serialize(wb)


def export_rows_xlsx(rows: list[tuple]) -> bytes:
    """Exporte une liste ``(batch, context_dict)`` en un classeur unique.

    ``context_dict`` provient de :func:`batch_context`.
    """
    wb, ws = _new_sheet()
    for batch, ctx in rows:
        ws.append(_row_values(batch, ctx))
    return _serialize(wb)


def export_packing_list_xlsx(batches, *, voyage_id, vessel, pol_code, pod_code) -> bytes:
    """Export d'une packing list : un batch = une ligne."""
    rows = [
        (
            b,
            batch_context(
                b, voyage_id=voyage_id, vessel=vessel, pol_code=pol_code, pod_code=pod_code
            ),
        )
        for b in batches
    ]
    return export_rows_xlsx(rows)


def parse_xlsx(content: bytes) -> list[dict]:
    """Relit un classeur → liste de dicts de valeurs typées (une par batch).

    Seules les colonnes éditables reconnues sont mappées ; les colonnes de
    contexte sont ignorées. Une ligne entièrement vide est sautée. Une cellule
    vide n'est pas reportée (le défaut de colonne s'applique à la création).
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    # En-tête → index ; on ne retient que les colonnes éditables connues.
    header_to_field = dict(EDITABLE_COLUMNS)
    col_field: dict[int, str] = {}
    for idx, name in enumerate(header_row):
        if name is None:
            continue
        field = header_to_field.get(str(name).strip().upper())
        if field is not None:
            col_field[idx] = field

    out: list[dict] = []
    for row in rows_iter:
        if row is None or not any(v not in (None, "") for v in row):
            continue
        form: dict = {}
        for idx, field in col_field.items():
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None or str(value).strip() == "":
                continue
            form[field] = value
        vals = {k: v for k, v in coerce_batch_form(form).items() if v is not None}
        if vals:
            out.append(vals)
    wb.close()
    return out
