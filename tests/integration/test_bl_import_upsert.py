"""L'import Excel met à jour au lieu de détruire — §4.4.

L'import faisait « delete-all + recreate ». Conséquence sur un registre de
connaissements : chaque import **détruisait les lots existants, donc leur
`bl_number`**. Autrement dit chaque import consommait des numéros de connaissement
sans retour, et cassait les liens déjà transmis au client.

Ce que ces tests protègent, par ordre de gravité :

1. **🔴 un lot numéroté survit à un import** — son numéro et l'état de son BL sont
   préservés ;
2. **un lot absent de l'import et déjà numéroté n'est pas supprimé**, et l'audit le
   **dit** plutôt que de laisser croire à une synchronisation complète ;
3. **la règle de régression s'applique** : une mise à jour effective annule la
   validation client, exactement comme une édition manuelle ;
4. **les deux portes se comportent pareil** — staff et portail expéditeur. Un
   garde-fou qui n'existe que d'un côté se contourne par l'autre.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.services import bl_workflow as w
from app.services import cargo_excel

TOKEN = "u" * 24


class _Req:
    headers: dict[str, str] = {}
    client = SimpleNamespace(host="203.0.113.31")
    url = SimpleNamespace(path="/cargo")
    state = SimpleNamespace(csrf_token="x")
    cookies: dict[str, str] = {}
    query_params: dict[str, str] = {}


class _Upload:
    def __init__(self, content: bytes, filename="import.xlsx"):
        self._content = content
        self.filename = filename

    async def read(self):
        return self._content


async def _ctx(db, *, count=2, with_token=False):
    from app.models.commercial import Client, Order
    from app.models.leg import Leg
    from app.models.packing_list import PackingList, PackingListBatch
    from app.models.port import Port
    from app.models.user import User
    from app.models.vessel import Vessel

    v = Vessel(name="Anemos", code="1")
    pol = Port(locode="FRFEC", name="Fécamp", country="FR")
    pod = Port(locode="BRSSO", name="Santos", country="BR")
    db.add_all([v, pol, pod])
    await db.flush()
    base = datetime(2026, 8, 10, tzinfo=UTC)
    leg = Leg(
        leg_code="1CFRBR6",
        vessel_id=v.id,
        departure_port_id=pol.id,
        arrival_port_id=pod.id,
        etd_ref=base,
        eta_ref=base + timedelta(days=20),
        etd=base,
        eta=base + timedelta(days=20),
    )
    cl = Client(name="ACME", client_type="shipper")
    db.add_all([leg, cl])
    await db.flush()
    order = Order(reference="CMD-UP", client_id=cl.id, leg_id=leg.id)
    db.add(order)
    await db.flush()
    pl = PackingList(
        order_id=order.id,
        leg_id=leg.id,
        token=TOKEN if with_token else None,
        token_expires_at=datetime.now(UTC) + timedelta(days=30) if with_token else None,
    )
    db.add(pl)
    await db.flush()
    ops = User(
        username="ops-up", email="ops-up@newtowt.test", hashed_password="x", role="operation"
    )
    db.add(ops)
    batches = []
    for i in range(1, count + 1):
        b = PackingListBatch(
            packing_list_id=pl.id,
            batch_number=i,
            pallet_format="EPAL",
            pallet_count=i * 2,
            consignee_name=f"Buyer {i}",
        )
        db.add(b)
        batches.append(b)
    await db.flush()
    return leg, pl, batches, ops


def _workbook(rows: list[dict]) -> bytes:
    """Classeur au gabarit officiel, une ligne par dict {en-tête: valeur}."""
    wb = load_workbook(BytesIO(cargo_excel.build_template_xlsx()))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for r, row in enumerate(rows, start=2):
        for header, value in row.items():
            ws.cell(row=r, column=headers.index(header) + 1, value=value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ───────────────── l'analyse remonte la clé de rapprochement ─────────────────


def test_the_parser_returns_the_match_key():
    content = _workbook([{"BATCH_NUMBER": 7, "PALLET_COUNT": 3}])
    parsed = cargo_excel.parse_xlsx(content)
    assert len(parsed) == 1
    assert parsed[0][cargo_excel.MATCH_KEY] == 7
    assert parsed[0]["pallet_count"] == 3


def test_an_unreadable_match_key_is_ignored_rather_than_guessed():
    """Un rapprochement hasardeux sur un registre de connaissements serait pire
    qu'une création."""
    content = _workbook([{"BATCH_NUMBER": "lot n°3", "PALLET_COUNT": 3}])
    parsed = cargo_excel.parse_xlsx(content)
    assert cargo_excel.MATCH_KEY not in parsed[0]


def test_the_match_key_is_never_confused_with_a_model_field():
    """Préfixée d'un souligné : l'appelant la retire avant d'écrire en base."""
    assert cargo_excel.MATCH_KEY.startswith("_")


# ───────────────── 🔴 un lot numéroté survit à un import ─────────────────


@pytest.mark.asyncio
async def test_a_numbered_batch_keeps_its_number_through_an_import(db):
    """🔴 Le cœur du correctif.

    Avant, l'import détruisait le lot : le numéro de connaissement était consommé pour
    rien et le lien déjà transmis au client cassé.
    """
    from app.models.packing_list import PackingListBatch
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, batches, ops = await _ctx(db)
    number = await w.generate_draft(db, pl=pl, batch=batches[0], leg=_leg, user=ops)
    batch_id = batches[0].id

    content = _workbook(
        [
            {"BATCH_NUMBER": 1, "PALLET_COUNT": 99, "CONSIGNEE_NAME": "Buyer 1 corrigé"},
            {"BATCH_NUMBER": 2, "PALLET_COUNT": 4},
        ]
    )
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    fresh = await db.get(PackingListBatch, batch_id)
    assert fresh is not None, "le lot a été détruit par l'import"
    assert fresh.bl_number == number, "le numéro de connaissement a été perdu"
    assert fresh.bl_state == w.DRAFT
    # …et la mise à jour a bien été appliquée.
    assert fresh.pallet_count == 99
    assert fresh.consignee_name == "Buyer 1 corrigé"


@pytest.mark.asyncio
async def test_a_row_without_a_match_key_creates_a_new_batch(db):
    from app.models.packing_list import PackingListBatch
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, _batches, ops = await _ctx(db, count=1)
    content = _workbook([{"PALLET_COUNT": 5, "CONSIGNEE_NAME": "Nouveau"}])
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    rows = list(
        (
            await db.execute(
                select(PackingListBatch).where(PackingListBatch.packing_list_id == pl.id)
            )
        )
        .scalars()
        .all()
    )
    assert any(b.consignee_name == "Nouveau" for b in rows)


# ───────── les absents de l'import : supprimés seulement si sans BL ─────────


@pytest.mark.asyncio
async def test_a_numberless_batch_absent_from_the_import_is_removed(db):
    """Le comportement « l'import fait foi » est conservé pour ce qui n'engage rien."""
    from app.models.packing_list import PackingListBatch
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, batches, ops = await _ctx(db, count=2)
    doomed_id = batches[1].id
    content = _workbook([{"BATCH_NUMBER": 1, "PALLET_COUNT": 2}])
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    assert await db.get(PackingListBatch, doomed_id) is None


@pytest.mark.asyncio
async def test_a_numbered_batch_absent_from_the_import_is_kept_and_said_so(db):
    """🔴 Supprimer un lot numéroté consommerait son numéro sans retour.

    Et le silence serait pire que la conservation : l'audit doit dire ce qui a été
    conservé, sinon on croit à une synchronisation complète.
    """
    from app.models.packing_list import PackingListAudit, PackingListBatch
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, batches, ops = await _ctx(db, count=2)
    number = await w.generate_draft(db, pl=pl, batch=batches[1], leg=_leg, user=ops)
    kept_id = batches[1].id

    content = _workbook([{"BATCH_NUMBER": 1, "PALLET_COUNT": 2}])  # le lot 2 est absent
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    assert await db.get(PackingListBatch, kept_id) is not None

    audits = list(
        (
            await db.execute(
                select(PackingListAudit).where(PackingListAudit.field == "_import_excel")
            )
        )
        .scalars()
        .all()
    )
    assert len(audits) == 1
    summary = audits[0].new_value or ""
    assert "conservés car déjà numérotés" in summary
    assert number in summary


@pytest.mark.asyncio
async def test_the_audit_summary_counts_updates_and_creations_separately(db):
    """« 3 batches importés » ne disait pas ce qui avait été détruit."""
    from app.models.packing_list import PackingListAudit
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, _batches, ops = await _ctx(db, count=2)
    content = _workbook(
        [
            {"BATCH_NUMBER": 1, "PALLET_COUNT": 7},
            {"BATCH_NUMBER": 2, "PALLET_COUNT": 8},
            {"PALLET_COUNT": 9},
        ]
    )
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    audit = (
        await db.execute(select(PackingListAudit).where(PackingListAudit.field == "_import_excel"))
    ).scalar_one()
    assert "2 mis à jour" in (audit.new_value or "")
    assert "1 créés" in (audit.new_value or "")


# ───────────────── la règle de régression s'applique ─────────────────


@pytest.mark.asyncio
async def test_an_import_that_changes_a_validated_batch_returns_it_to_draft(db):
    """Une validation porte sur un contenu : un import qui le modifie l'annule."""
    from app.models.client_account import ClientAccount
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, batches, ops = await _ctx(db, count=1)
    account = ClientAccount(email="c-up@example.test", hashed_password="x", company_name="Belco")
    db.add(account)
    await db.flush()
    await w.generate_draft(db, pl=pl, batch=batches[0], leg=_leg, user=ops)
    await w.validate_by_client(db, batch=batches[0], client=account)
    assert batches[0].bl_state == w.CLIENT_VALIDATED

    content = _workbook([{"BATCH_NUMBER": 1, "PALLET_COUNT": 42}])
    await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)

    assert batches[0].bl_state == w.DRAFT, "la validation client a survécu à une modification"


@pytest.mark.asyncio
async def test_a_signed_batch_still_blocks_the_whole_import(db):
    """⚠️ Garde anti-régression : le refus en bloc du lot précédent reste en place.

    L'upsert préserve les numéros, mais un lot **signé** ne doit pas même être
    modifié — la correction passe par une révision.
    """
    from fastapi import HTTPException

    from app.models.client_account import ClientAccount
    from app.models.user import User
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    _leg, pl, batches, ops = await _ctx(db, count=1)
    account = ClientAccount(email="c-up2@example.test", hashed_password="x", company_name="Belco")
    master = User(
        username="cdt-up", email="cdt-up@newtowt.test", hashed_password="x", role="marins"
    )
    db.add_all([account, master])
    await db.flush()
    await w.generate_draft(db, pl=pl, batch=batches[0], leg=_leg, user=ops)
    await w.validate_by_client(db, batch=batches[0], client=account)
    await w.sign_by_master(db, batch=batches[0], user=master)

    content = _workbook([{"BATCH_NUMBER": 1, "PALLET_COUNT": 42}])
    with pytest.raises(HTTPException) as e:
        await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=ops)
    assert e.value.status_code == 409
    assert batches[0].pallet_count != 42


# ───────────────── les deux portes se comportent pareil ─────────────────


@pytest.mark.asyncio
async def test_the_portal_import_also_preserves_numbered_batches(db):
    """Un garde-fou qui n'existe que d'un côté se contourne par l'autre."""
    from app.models.packing_list import PackingListBatch
    from app.routers.cargo_portal_router import portal_packing_import_xlsx

    _leg, pl, batches, ops = await _ctx(db, count=2, with_token=True)
    number = await w.generate_draft(db, pl=pl, batch=batches[0], leg=_leg, user=ops)
    batch_id = batches[0].id

    content = _workbook([{"BATCH_NUMBER": 1, "PALLET_COUNT": 55}])
    await portal_packing_import_xlsx(TOKEN, _Req(), file=_Upload(content), db=db)

    fresh = await db.get(PackingListBatch, batch_id)
    assert fresh is not None and fresh.bl_number == number
    assert fresh.pallet_count == 55
