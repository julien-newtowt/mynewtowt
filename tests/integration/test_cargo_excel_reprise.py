"""Cargo P1 — reprise (CARGO-09 import/export Excel de la packing list)."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.commercial import Client, Order
from app.models.leg import Leg
from app.models.packing_list import PackingList, PackingListBatch
from app.models.port import Port
from app.models.vessel import Vessel
from app.services import cargo_excel


class _Req:
    headers: dict[str, str] = {}
    client = SimpleNamespace(host="127.0.0.1")
    url = SimpleNamespace(path="/p/x/packing/import-xlsx")


class _Upload:
    def __init__(self, content: bytes, filename: str = "import.xlsx"):
        self.filename = filename
        self._content = content

    async def read(self) -> bytes:
        return self._content


async def _pl_with_batches(db, *, with_leg: bool = True):
    c = Client(name="ACME", client_type="shipper")
    db.add(c)
    if with_leg:
        db.add(Vessel(id=1, code="ANE", name="Anemos"))
        db.add(Port(id=1, locode="FRFEC", name="Fécamp", country="FR"))
        db.add(Port(id=2, locode="BRSSO", name="Santos", country="BR"))
        await db.flush()
        base = datetime(2026, 4, 1, tzinfo=UTC)
        db.add(
            Leg(
                id=1,
                leg_code="1CFRBR6",
                vessel_id=1,
                departure_port_id=1,
                arrival_port_id=2,
                etd_ref=base,
                eta_ref=base + timedelta(days=20),
                etd=base,
                eta=base + timedelta(days=20),
            )
        )
    await db.flush()
    o = Order(reference="ORD-2026-0001", client_id=c.id, leg_id=1 if with_leg else None)
    db.add(o)
    await db.flush()
    pl = PackingList(order_id=o.id, status="draft")
    db.add(pl)
    await db.flush()
    db.add(
        PackingListBatch(
            packing_list_id=pl.id,
            batch_number=1,
            pallet_format="USPAL",
            pallet_count=12,
            type_of_goods="Vin",
            weight_kg=480.0,
            cases_quantity=24,
            hazardous=True,
            stackable=False,
            shipper_name="Domaine X",
        )
    )
    await db.flush()
    return pl


# ─────────────────────────── service round-trip ───────────────────────────


def test_template_has_all_headers():
    wb = load_workbook(BytesIO(cargo_excel.build_template_xlsx()))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == list(cargo_excel.ALL_HEADERS)
    # pas de ligne de données
    assert ws.max_row == 1


def test_export_then_parse_roundtrip():
    b = PackingListBatch(
        packing_list_id=1,
        batch_number=1,
        pallet_format="IBC",
        pallet_count=8,
        type_of_goods="Cognac",
        weight_kg=300.0,
        cases_quantity=10,
        units_per_case=6,
        cargo_value_usd=1500.0,
        hazardous=True,
        stackable=False,
        shipper_name="Shipper SA",
        consignee_name="Cons SA",
    )
    content = cargo_excel.export_packing_list_xlsx(
        [b], voyage_id="1CFRBR6", vessel="Anemos", pol_code="FRFEC", pod_code="BRSSO"
    )
    parsed = cargo_excel.parse_xlsx(content)
    assert len(parsed) == 1
    vals = parsed[0]
    assert vals["pallet_format"] == "IBC"
    assert vals["pallet_count"] == 8
    assert vals["cases_quantity"] == 10
    assert vals["units_per_case"] == 6
    assert vals["cargo_value_usd"] == 1500.0
    assert vals["hazardous"] is True
    assert vals["stackable"] is False
    assert vals["shipper_name"] == "Shipper SA"
    # colonnes de contexte ignorées à l'import
    assert "voyage_id" not in vals and "VOYAGE_ID" not in vals


def test_parse_skips_empty_rows_and_unknown_columns():
    wb = load_workbook(BytesIO(cargo_excel.build_template_xlsx()))
    ws = wb.active
    # ligne vide → sautée ; une ligne avec seulement PALLET_COUNT
    headers = [c.value for c in ws[1]]
    idx_count = headers.index("PALLET_COUNT") + 1
    ws.cell(row=3, column=idx_count, value=5)
    buf = BytesIO()
    wb.save(buf)
    parsed = cargo_excel.parse_xlsx(buf.getvalue())
    assert len(parsed) == 1
    assert parsed[0]["pallet_count"] == 5
    # stackable absent → pas dans vals (défaut de colonne s'applique)
    assert "stackable" not in parsed[0]


# ─────────────────────────── routes staff ───────────────────────────


@pytest.mark.asyncio
async def test_staff_export_xlsx(db, staff_user):
    from app.routers.cargo_packing_router import packing_list_export_xlsx

    pl = await _pl_with_batches(db)
    resp = await packing_list_export_xlsx(pl.id, db=db, user=staff_user)
    assert resp.media_type == cargo_excel.XLSX_MIME
    ws = load_workbook(BytesIO(resp.body)).active
    assert ws.max_row == 2  # en-tête + 1 batch


@pytest.mark.asyncio
async def test_staff_import_replaces_batches(db, staff_user):
    from app.routers.cargo_packing_router import packing_list_import_xlsx

    pl = await _pl_with_batches(db)
    # nouveau classeur avec 2 batches
    content = cargo_excel.export_packing_list_xlsx(
        [
            PackingListBatch(
                packing_list_id=pl.id, batch_number=1, pallet_count=3, type_of_goods="A"
            ),
            PackingListBatch(
                packing_list_id=pl.id, batch_number=2, pallet_count=7, type_of_goods="B"
            ),
        ],
        voyage_id=None,
        vessel=None,
        pol_code=None,
        pod_code=None,
    )
    resp = await packing_list_import_xlsx(
        pl.id, _Req(), file=_Upload(content), db=db, user=staff_user
    )
    assert resp.status_code == 303
    batches = (
        (
            await db.execute(
                select(PackingListBatch).where(PackingListBatch.packing_list_id == pl.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(batches) == 2
    assert sorted(b.pallet_count for b in batches) == [3, 7]
    assert sorted(b.batch_number for b in batches) == [1, 2]


@pytest.mark.asyncio
async def test_staff_import_rejected_when_locked(db, staff_user):
    from fastapi import HTTPException

    from app.routers.cargo_packing_router import packing_list_import_xlsx

    pl = await _pl_with_batches(db)
    pl.status = "locked"
    await db.flush()
    content = cargo_excel.build_template_xlsx()
    with pytest.raises(HTTPException) as exc:
        await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=staff_user)
    assert exc.value.status_code == 409


@pytest.mark.asyncio
async def test_staff_import_empty_rejected(db, staff_user):
    from fastapi import HTTPException

    from app.routers.cargo_packing_router import packing_list_import_xlsx

    pl = await _pl_with_batches(db)
    # template vierge → 0 ligne exploitable → 400 (et batches existants préservés)
    content = cargo_excel.build_template_xlsx()
    with pytest.raises(HTTPException) as exc:
        await packing_list_import_xlsx(pl.id, _Req(), file=_Upload(content), db=db, user=staff_user)
    assert exc.value.status_code == 400
    batches = (
        (
            await db.execute(
                select(PackingListBatch).where(PackingListBatch.packing_list_id == pl.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(batches) == 1  # rien n'a été supprimé


@pytest.mark.asyncio
async def test_staff_import_oversize_rejected(db, staff_user):
    """Un fichier > 20 Mo est rejeté (413) même sans Content-Length fiable."""
    from fastapi import HTTPException

    from app.routers.cargo_packing_router import packing_list_import_xlsx

    pl = await _pl_with_batches(db)
    oversized = b"\x00" * (21 * 1024 * 1024)
    with pytest.raises(HTTPException) as exc:
        await packing_list_import_xlsx(
            pl.id, _Req(), file=_Upload(oversized), db=db, user=staff_user
        )
    assert exc.value.status_code == 413


@pytest.mark.asyncio
async def test_voyage_export_aggregates_all_pls(db, staff_user):
    from app.routers.cargo_packing_router import voyage_export_xlsx

    pl = await _pl_with_batches(db)
    # 2e PL sur le même leg (réutilise le client de la 1re commande)
    client_id = (await db.get(Order, pl.order_id)).client_id
    o2 = Order(reference="ORD-2026-0002", client_id=client_id, leg_id=1)
    db.add(o2)
    await db.flush()
    pl2 = PackingList(order_id=o2.id, status="draft")
    db.add(pl2)
    await db.flush()
    db.add(PackingListBatch(packing_list_id=pl2.id, batch_number=1, pallet_count=2))
    await db.flush()

    resp = await voyage_export_xlsx(1, db=db, user=staff_user)
    ws = load_workbook(BytesIO(resp.body)).active
    assert ws.max_row == 3  # en-tête + 2 batches (1 par PL)


# ─────────────────────────── routes portail ───────────────────────────


@pytest.mark.asyncio
async def test_portal_import_replaces_batches(db):
    from app.routers.cargo_portal_router import portal_packing_import_xlsx

    pl = await _pl_with_batches(db)
    content = cargo_excel.export_packing_list_xlsx(
        [PackingListBatch(packing_list_id=pl.id, batch_number=1, pallet_count=9)],
        voyage_id=None,
        vessel=None,
        pol_code=None,
        pod_code=None,
    )
    resp = await portal_packing_import_xlsx(pl.token, _Req(), file=_Upload(content), db=db)
    assert resp.status_code == 303
    batches = (
        (
            await db.execute(
                select(PackingListBatch).where(PackingListBatch.packing_list_id == pl.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(batches) == 1 and batches[0].pallet_count == 9
