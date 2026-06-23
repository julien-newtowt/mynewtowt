"""ADM-05 — reprise de l'import en masse d'utilisateurs (Excel)."""

from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.user import User
from app.services import user_import


class _Req:
    headers: dict[str, str] = {}
    cookies: dict[str, str] = {}
    query_params: dict[str, str] = {}
    client = SimpleNamespace(host="127.0.0.1")
    url = SimpleNamespace(path="/admin/users/import")
    state = SimpleNamespace(
        notif_count=0,
        newtowt_agent_enabled=True,
        recent_notifications=[],
        csrf_token="test-csrf",
    )


class _Upload:
    def __init__(self, content: bytes):
        self.filename = "users.xlsx"
        self._content = content

    async def read(self) -> bytes:
        return self._content


def _book(rows: list[dict]) -> bytes:
    """Construit un classeur d'import à partir de dicts."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = list(user_import.IMPORT_COLUMNS)
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_has_columns():
    ws = load_workbook(BytesIO(user_import.build_template_xlsx())).active
    assert [c.value for c in ws[1]] == list(user_import.IMPORT_COLUMNS)


def test_parse_reads_rows():
    content = _book([{"username": "jdoe", "email": "j@x.io", "role": "operation"}])
    rows = user_import.parse_users_xlsx(content)
    assert rows == [{"username": "jdoe", "email": "j@x.io", "role": "operation"}]


@pytest.mark.asyncio
async def test_import_creates_valid_and_reports_invalid(db):
    content = _book(
        [
            {"username": "jdoe", "email": "j@x.io", "full_name": "Jane Doe", "role": "operation"},
            {"username": "bad", "email": "not-an-email", "role": "operation"},  # email KO
            {"username": "norole", "email": "n@x.io", "role": "wizard"},  # rôle KO
            {"username": "jdoe", "email": "dup@x.io", "role": "operation"},  # doublon username
            {},  # ligne vide → sautée au parse
        ]
    )
    rows = user_import.parse_users_xlsx(content)
    assert len(rows) == 4  # la ligne vide est éliminée au parse
    report = await user_import.import_users(db, rows)
    assert len(report["created"]) == 1
    assert report["created"][0]["username"] == "jdoe"
    assert report["created"][0]["temp_password"]  # généré
    assert len(report["skipped"]) == 3  # email, rôle, doublon

    u = (await db.execute(select(User).where(User.username == "jdoe"))).scalar_one()
    assert u.must_change_password is True
    assert u.role == "operation"
    assert u.email == "j@x.io"


@pytest.mark.asyncio
async def test_import_skips_existing(db):
    db.add(
        User(
            username="exists",
            email="e@x.io",
            hashed_password="x",
            role="operation",
        )
    )
    await db.flush()
    content = _book([{"username": "exists", "email": "e@x.io", "role": "operation"}])
    report = await user_import.import_users(db, user_import.parse_users_xlsx(content))
    assert report["created"] == []
    assert report["skipped"][0]["reason"] == "utilisateur déjà existant"


@pytest.mark.asyncio
async def test_import_route_renders_report(db, staff_user):
    from app.routers.admin_router import users_import

    content = _book([{"username": "jdoe", "email": "j@x.io", "role": "operation"}])
    resp = await users_import(_Req(), file=_Upload(content), db=db, user=staff_user)
    assert resp.status_code == 200
    assert resp.template.name == "staff/admin/users_import.html"
    assert len(resp.context["report"]["created"]) == 1


def test_import_templates_compile():
    from app.templating import templates

    for n in ("staff/admin/users_import.html", "staff/admin/users.html"):
        assert templates.env.get_template(n) is not None
