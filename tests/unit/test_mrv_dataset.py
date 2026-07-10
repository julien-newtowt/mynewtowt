"""Tests — sorties réglementaires OVDLA / OVDBR (LOT 10).

Couvre ``app.services.mrv_dataset`` sur une chaîne synthétique de référence
(Departure + 2 Noon + Arrival), moteur SQLite en mémoire (FK activées, seed du
référentiel de validation → densité R16 = 0,845, moteurs ME/AE) :

- **deltas** : la ligne Arrival porte la ΣΔ des 3 intervalles (agrégation des
  intervalles noon) ; les Noon ne produisent PAS de ligne (Q10) ;
- **DMS** exact (décimal → degrés + minutes entières) ;
- **Period last event** (clôture de période, voyage ouvert au period_end) ;
- **portes** : événement non validé exclu ; événement lié à un rapport
  ``under_conformity`` exclu + alerte admin ;
- **OVDBR** : mapping BDN, porte ``valide_master`` ;
- **snapshot** upsert idempotent (``verification_status`` préservé) ;
- **export xlsx** relisible par openpyxl aux mêmes valeurs.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime, timedelta
from decimal import Decimal

import openpyxl
import pytest_asyncio
from sqlalchemy import event, select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401
from app.database import Base
from app.models.bunker import BunkerOperation
from app.models.env_report import EnvFieldModification, EnvReport, EnvReportEventLink
from app.models.leg import Leg
from app.models.mrv_dataset import MrvLogAbstractEntry
from app.models.nav_event import (
    ArrivalEvent,
    DepartureEvent,
    NavEventEngineReading,
    NoonEvent,
)
from app.models.notification import Notification
from app.models.port import Port
from app.models.vessel import Vessel
from app.services import mrv_dataset as md
from app.services.referential_env import ensure_vessel_env_defaults, get_vessel_engines
from app.services.validation_engine import invalidate_cache, seed_reference_data

FACTOR = Decimal("0.001") * Decimal("0.845")
T0 = datetime(2026, 1, 1, 12, 0, tzinfo=UTC)

# Compteurs carburant (litres bruts) par moteur / événement de la chaîne.
DEP_FUEL = {"PME": 10000, "SME": 8000, "FWD_GEN": 5000, "AFT_GEN": 4000}
N1_FUEL = {"PME": 11000, "SME": 8600, "FWD_GEN": 5300, "AFT_GEN": 4200}
N2_FUEL = {"PME": 12000, "SME": 9200, "FWD_GEN": 5600, "AFT_GEN": 4400}
ARR_FUEL = {"PME": 12500, "SME": 9500, "FWD_GEN": 5750, "AFT_GEN": 4500}


@pytest_asyncio.fixture
async def db():
    engine = create_async_engine(
        "sqlite+aiosqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk(dbapi_conn, _rec):  # pragma: no cover
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session = async_sessionmaker(engine, expire_on_commit=False)()
    invalidate_cache()
    await seed_reference_data(session)
    invalidate_cache()
    try:
        yield session
    finally:
        await session.close()
        await engine.dispose()
        invalidate_cache()


def _readings(engines, fuel_map):
    return [
        NavEventEngineReading(
            engine_id=engines[r].id, fuel_counter_l=Decimal(str(v)), is_counter_reset=False
        )
        for r, v in fuel_map.items()
    ]


async def _chain(db, *, arrival_status: str = "valide"):
    """Chaîne D + 2N + A validée (positions Fécamp→Belém), + engines."""
    vessel = Vessel(code="ANE", name="Anemos", imo_number="9982938")
    db.add(vessel)
    await db.flush()
    await ensure_vessel_env_defaults(db, vessel)
    engines = {e.engine_role: e for e in await get_vessel_engines(db, vessel.id)}
    p1 = Port(name="Fecamp", country="FR", locode="FRFEC", latitude=49.7, longitude=0.37)
    p2 = Port(name="Belem", country="BR", locode="BRBEL", latitude=-1.45, longitude=-48.5)
    db.add_all([p1, p2])
    await db.flush()
    leg = Leg(
        leg_code="1AFRBR6",
        vessel_id=vessel.id,
        departure_port_id=p1.id,
        arrival_port_id=p2.id,
        etd_ref=T0,
        eta_ref=T0 + timedelta(days=3),
        etd=T0,
        eta=T0 + timedelta(days=3),
    )
    db.add(leg)
    await db.flush()

    dep = DepartureEvent(
        leg_id=leg.id,
        vessel_id=vessel.id,
        status="valide",
        datetime_utc=T0,
        lat_decimal=Decimal("47.8167"),
        lon_decimal=Decimal("-3.9333"),
        rob_t=Decimal("100.000"),
        vessel_condition="laden",
        cargo_bl_t=Decimal("500.000"),
        cargo_mrv_t=Decimal("540.000"),
    )
    dep.engine_readings = _readings(engines, DEP_FUEL)
    n1 = NoonEvent(
        leg_id=leg.id,
        vessel_id=vessel.id,
        status="valide",
        datetime_utc=T0 + timedelta(hours=24),
        lat_decimal=Decimal("45.0"),
        lon_decimal=Decimal("-10.0"),
    )
    n1.engine_readings = _readings(engines, N1_FUEL)
    n2 = NoonEvent(
        leg_id=leg.id,
        vessel_id=vessel.id,
        status="valide",
        datetime_utc=T0 + timedelta(hours=48),
        lat_decimal=Decimal("40.0"),
        lon_decimal=Decimal("-20.0"),
    )
    n2.engine_readings = _readings(engines, N2_FUEL)
    arr = ArrivalEvent(
        leg_id=leg.id,
        vessel_id=vessel.id,
        status=arrival_status,
        datetime_utc=T0 + timedelta(hours=72),
        lat_decimal=Decimal("35.0"),
        lon_decimal=Decimal("-30.0"),
        rob_t=Decimal("96.000"),
        vessel_condition="laden",
        cargo_mrv_t=Decimal("540.000"),
    )
    arr.engine_readings = _readings(engines, ARR_FUEL)
    db.add_all([dep, n1, n2, arr])
    await db.flush()
    return vessel, leg, dep, n1, n2, arr


# ═══════════════════════════════════════════════ deltas + exclusion des Noon


async def test_ovdla_arrival_aggregates_three_intervals(db):
    vessel, leg, dep, n1, n2, arr = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)

    # Pas de ligne Noon (Q10) : seulement Departure + Arrival.
    assert [r.values["Event"] for r in rows] == ["Departure", "Arrival"]
    assert all(r.included for r in rows)

    arr_row = rows[1].values
    # ME ΔL total = (12500-10000)+(9500-8000)=4000 → 4000×0,001×0,845 = 3,38 t
    # AE ΔL total = (5750-5000)+(4500-4000)=1250 → 1,05625 t
    assert arr_row["ME_Consumption_MDO"] == Decimal("4000") * FACTOR
    assert arr_row["AE_Consumption_MDO"] == Decimal("1250") * FACTOR
    # Temps depuis la ligne OVDLA précédente (Departure) = 72 h.
    assert arr_row["Time_Since_Previous_Report"] == Decimal("72.000")
    # Distance = polyline haversine des 3 intervalles (> distance directe D→A).
    assert arr_row["Distance"] > 0
    # ROB déclaré (source R14-v2) reproduit exactement.
    assert arr_row["MDO_ROB"] == Decimal("96.000")
    # Cargo + ports.
    assert arr_row["Cargo_Mt"] == Decimal("540.000")
    assert (arr_row["Voyage_From"], arr_row["Voyage_To"]) == ("FRFEC", "BRBEL")
    # Departure (1re ligne) : pas d'intervalle amont → conso nulle, ROB déclaré.
    assert rows[0].values["ME_Consumption_MDO"] == Decimal("0.00000")
    assert rows[0].values["MDO_ROB"] == Decimal("100.000")


async def test_ovdla_dms_exact(db):
    vessel, *_ = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)
    dep = rows[0].values
    # 47,8167 → 47°49'N ; -3,9333 → 3°56'W (minutes ENTIÈRES, cf. échantillons).
    assert (dep["Latitude_North_South"], dep["Latitude_Degree"], dep["Latitude_Minutes"]) == (
        "N",
        47,
        49,
    )
    assert (dep["Longitude_East_West"], dep["Longitude_Degree"], dep["Longitude_Minutes"]) == (
        "W",
        3,
        56,
    )


async def test_ovdla_source_system_is_mytowt(db):
    vessel, *_ = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)
    assert all(r.values["Source_System"] == "MyTOWT" for r in rows)


# ═══════════════════════════════════════════════ Period last event (Q10)


async def test_period_last_event_on_open_voyage(db):
    vessel, leg, dep, n1, n2, arr = await _chain(db)
    # Voyage « ouvert » : on coupe la période AVANT l'Arrival (des Noon restent
    # après la dernière ligne OVDLA jusqu'à period_end) → ligne synthétique.
    period_end = T0 + timedelta(hours=60)  # après N2 (48 h), avant Arrival (72 h)
    rows = await md.build_ovdla_rows(db, vessel, T0 - timedelta(hours=1), period_end)
    labels = [r.values["Event"] for r in rows]
    assert "Period last event" in labels
    marker = next(r for r in rows if r.values["Event"] == "Period last event")
    assert marker.synthetic is True
    assert marker.event_id is None  # non persistée (aucun événement rattaché)
    # Datée à la clôture de période, conso agrégée depuis la dernière ligne (Departure).
    assert marker.values["Date_UTC"] == period_end.date()
    assert marker.values["ME_Consumption_MDO"] > 0


# ═══════════════════════════════════════════════ Portes de génération


async def test_gate_non_validated_event_excluded(db):
    vessel, leg, dep, n1, n2, arr = await _chain(db, arrival_status="finalise")
    rows = await md.build_ovdla_rows(db, vessel)
    arr_row = next(r for r in rows if r.values["Event"] == "Arrival")
    assert arr_row.included is False
    assert "non validé" in (arr_row.exclusion_reason or "")


async def test_gate_under_conformity_excluded_and_alerts(db):
    vessel, leg, dep, n1, n2, arr = await _chain(db)
    # Rapport lié à l'Arrival avec une modification under_conformity.
    report = EnvReport(leg_id=leg.id, report_type="carbon", status="valide_master", payload={})
    db.add(report)
    await db.flush()
    db.add(EnvReportEventLink(report_id=report.id, event_id=arr.id))
    db.add(
        EnvFieldModification(
            report_id=report.id,
            field_name="rob_t",
            justification_text="écart",
            resulting_quality_status="under_conformity",
        )
    )
    await db.flush()

    rows = await md.build_ovdla_rows(db, vessel, alert=True)
    arr_row = next(r for r in rows if r.values["Event"] == "Arrival")
    assert arr_row.included is False
    assert arr_row.verification_status == "under_conformity"
    assert "under_conformity" in (arr_row.exclusion_reason or "")

    # Alerte admin émise (pattern lot 8).
    notifs = (
        (await db.execute(select(Notification).where(Notification.target_role == "administrateur")))
        .scalars()
        .all()
    )
    assert any("OVDLA" in (n.title or "") for n in notifs)


# ═══════════════════════════════════════════════ OVDBR


async def _bunker(db, vessel, leg, bdn, status, mass="30.000", when=None):
    b = BunkerOperation(
        leg_id=leg.id,
        vessel_id=vessel.id,
        bdn_number=bdn,
        port_locode="FRFEC",
        delivery_datetime_utc=(when or (T0 - timedelta(days=1))),
        fuel_type="MDO",
        mass_t=Decimal(mass),
        density_15c_t_m3=Decimal("0.845"),
        status=status,
    )
    db.add(b)
    await db.flush()
    return b


async def test_ovdbr_mapping_and_gate(db):
    vessel, leg, *_ = await _chain(db)
    await _bunker(db, vessel, leg, "433421", "valide_master", mass="30.054")
    await _bunker(db, vessel, leg, "DRAFT-1", "brouillon", mass="12.000")

    rows = await md.build_ovdbr_rows(db, vessel)
    ok = next(r for r in rows if r.values["BDN_Number"] == "433421")
    assert ok.included is True
    assert ok.values["Mass"] == Decimal("30.054")
    assert ok.values["Bunker_Port"] == "FRFEC"
    assert ok.values["Fuel_Type"] == "MDO"
    assert ok.values["Source_System"] == "MyTOWT"
    assert ok.values["Bunker_Delivery_Time"] is not None

    draft = next(r for r in rows if r.values["BDN_Number"] == "DRAFT-1")
    assert draft.included is False
    assert "validé Master" in (draft.exclusion_reason or "")


# ═══════════════════════════════════════════════ Snapshot idempotent


async def test_snapshot_idempotent_preserves_verification_status(db):
    vessel, leg, dep, n1, n2, arr = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)

    first = await md.snapshot_entries(db, rows)
    assert first["created"] == 2 and first["updated"] == 0  # Departure + Arrival

    # Une vérification manuelle fige le statut d'une entrée…
    entry = (
        await db.execute(select(MrvLogAbstractEntry).where(MrvLogAbstractEntry.event_id == arr.id))
    ).scalar_one()
    entry.verification_status = "corrected"
    await db.flush()

    # …une régénération met à jour le payload SANS écraser le statut.
    rows2 = await md.build_ovdla_rows(db, vessel)
    second = await md.snapshot_entries(db, rows2)
    assert second["created"] == 0 and second["updated"] == 2

    all_entries = (await db.execute(select(MrvLogAbstractEntry))).scalars().all()
    assert len(all_entries) == 2  # pas de doublon (event_id UNIQUE)
    refreshed = (
        await db.execute(select(MrvLogAbstractEntry).where(MrvLogAbstractEntry.event_id == arr.id))
    ).scalar_one()
    assert refreshed.verification_status == "corrected"


# ═══════════════════════════════════════════════ Export xlsx


async def test_export_xlsx_roundtrip(db):
    vessel, *_ = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)
    content = md.export_xlsx(rows, kind="ovdla")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == list(md.OVDLA_COLUMNS)
    # 1 en-tête + 2 lignes incluses (Departure, Arrival).
    assert ws.max_row == 3
    # La valeur ME de l'Arrival est relisible et égale à la valeur calculée.
    me_col = header.index("ME_Consumption_MDO") + 1
    arr_me = ws.cell(row=3, column=me_col).value
    assert abs(float(arr_me) - float(Decimal("4000") * FACTOR)) < 1e-6


async def test_export_csv_headers_and_rows(db):
    vessel, *_ = await _chain(db)
    rows = await md.build_ovdla_rows(db, vessel)
    csv_text = md.export_csv(rows, kind="ovdla")
    lines = csv_text.splitlines()
    assert lines[0].split(",") == list(md.OVDLA_COLUMNS)
    assert len([ln for ln in lines if ln]) == 3  # header + 2 rows


# ═══════════════════════════════════════════════ Anti-injection de formule


async def test_export_neutralises_formula_injection(db):
    """Un BDN texte commençant par ``=`` ne doit JAMAIS ressortir en formule
    vive dans le fichier déposé chez DNV — ni en CSV ni en XLSX (openpyxl
    classerait sinon la chaîne en ``data_type='f'``)."""
    import csv as _csv

    vessel, leg, *_ = await _chain(db)
    payload = '=HYPERLINK("http://evil")'  # 25 car. < String(40)
    await _bunker(db, vessel, leg, payload, "valide_master", mass="30.000")
    rows = await md.build_ovdbr_rows(db, vessel)
    bdn_idx = list(md.OVDBR_COLUMNS).index("BDN_Number")

    # CSV : cellule préfixée d'une apostrophe (neutralisée).
    data_row = list(_csv.reader(md.export_csv(rows, kind="ovdbr").splitlines()))[1]
    assert data_row[bdn_idx] == "'" + payload

    # XLSX : chaîne littérale (data_type 's'), jamais une formule ('f').
    wb = openpyxl.load_workbook(io.BytesIO(md.export_xlsx(rows, kind="ovdbr")), data_only=True)
    cell = wb.active.cell(row=2, column=bdn_idx + 1)
    assert cell.data_type == "s"
    assert str(cell.value).startswith("'")

    # Un nombre négatif (Mass) ne doit PAS être corrompu par la neutralisation.
    mass_idx = list(md.OVDBR_COLUMNS).index("Mass")
    assert data_row[mass_idx] == "30.000"
